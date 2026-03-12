import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from MatavimoVieta import Kaminas

def irasyti_antrastes(kaminas_obj: Kaminas):
    failas = "Rezultatai.xlsx"
    n_lin = kaminas_obj.liniju_skaicius
    
    # 1. Grąžinti visi pilni jūsų nurodyti pavadinimai
    stulpeliai = [
        "Matavimo data",
        "Ėminių registracijos Nr. T-107-2026-E-",
        "Objekto pavadinimas, adresas, taršos šaltinio Nr.",
        "Matavimo taškai ortakyje nuo vidinės sienelės, cm"
    ]
    
    # Diferencinis slėgis (dinamiškas kiekis)
    for i in range(1, n_lin + 1):
        stulpeliai.append(f"Išmatuotas diferencinis slėgis taškuose {i} linija, Pdi, hPa")
        
    stulpeliai.extend([
        "Pito vamzdelio koeficientas, K", 
        "Temperatūra ortakyje t, oC", 
        "Atmosferinis slėgis P, hPa",
        "Statinis slėgis ortakyje ± ΔP, hPa", 
        "Dujų slėgis kamine Pk= P+ ΔP, hPa", 
        "Dujų mol. masė Ms= qn *22.4, kg/"
    ])
    
    # Srauto greitis (dinamiškas kiekis)
    for i in range(1, n_lin + 1):
        stulpeliai.append(f"Dujų srauto greitis matavimo taškuose, {i} linija, wi = K*129 *√t *√Pdi/ √Pk/√Ms, m/s")
        
    stulpeliai.extend([
        "Vidutinis dujų srauto greitis ortakyje w, m/s", 
        "Ortakio diametras, matmenys, m", 
        "Ortakio skerspjūvio plotas F, m2",
        "Dujų tūrio debitas realiomis sąlygomis Vk = wvid × F, m3/s", 
        "Dujų tūrio debitas normaliosiomis sąlygomis Vdr n.s =Vk *0,269*(P±ΔP)/(273+t), m3/s",
        "Vandens kondensato masė m H2O, kg", 
        "Vandens garų konc. dujose x=mH2O/Vmn*qn, kg/kg",
        "Sausų dujų tūrio debitas normaliosiomis sąlygomis Vn= Vdr n.s *((1/(1+(x*qn/0.8038)))",
        "Prasiurbtas dujų tūris normaliosiomis sąlygomis Vmn, Nm3", 
        "Išplėstinės neapibrėžties koef. A",
        "Išplėstinė neapibrėžtis (dujų tūrio debitas) Uv", 
        "Išplėstinė neapibrėžtis (dujų srauto greitis) Uw",
        "Sausų dujų tankis normaliosioms sąlygomis qn, (kg/m3)", 
        "Skaičiavimus atliko (inicialai, data)"
    ])

    wb = load_workbook(failas)
    ws = wb["Greitis"]
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # 2. Antraštė 3-oje eilutėje su dinaminiu suliejimu
    paskutinis_stulpelis = len(stulpeliai)
    ws.cell(row=3, column=1, value="Debitas pagal matuotą diferencinį slėgį")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=paskutinis_stulpelis)
    
    pavadinimo_cell = ws.cell(row=3, column=1)
    pavadinimo_cell.font = Font(bold=True, size=12) # Galite padidinti šriftą, jei reikia
    pavadinimo_cell.alignment = Alignment(horizontal="center", vertical="center")

    # 3. Stulpelių formavimas (5 eilutė)
    for idx, pavadinimas in enumerate(stulpeliai, start=1):
        cell = ws.cell(row=5, column=idx, value=pavadinimas)
        cell.font = Font(bold=False) # Nuimtas paryškinimas
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
        # Stulpelio plotis - parinktas 15, kad būtų kompaktiškiau, bet tilptų tekstas
        col_letter = cell.column_letter
        ws.column_dimensions[col_letter].width = 15

    # 5 eilutės aukštis (didiname, nes pavadinimai ilgi ir turi talpintis langelyje)
    ws.row_dimensions[5].height = 130 

    # 4. Matmenų įrašymas po stulpeliu "Ortakio diametras, matmenys, m"
    
    for idx, pavadinimas in enumerate(stulpeliai, start=1):
        if "Ortakio diametras, matmenys, m" in pavadinimas:
            # Sukuriame stilių reikšmėms: centravimas ir rėmeliai
            centered_alignment = Alignment(horizontal="center", vertical="center")
            
            if kaminas_obj.forma == 'A':
                # Apvalus: reikšmė į pirmą langelį po fraze (6 eilutė)
                reiksme = f"{(kaminas_obj.skersmuo / 100):.2f}".replace('.', ',')
                cell = ws.cell(row=6, column=idx, value=reiksme)
                cell.alignment = centered_alignment
            else:
                # Stačiakampis: gylis į antrą (6 eilutė), plotis į trečią (7 eilutė)
                g_m = f"{(kaminas_obj.gylis / 100):.2f}".replace('.', ',')
                p_m = f"{(kaminas_obj.plotis / 100):.2f}".replace('.', ',')
                
                # Gylis
                cell_g = ws.cell(row=7, column=idx, value=g_m)
                cell_g.alignment = centered_alignment
                
                # Plotis
                cell_p = ws.cell(row=8, column=idx, value=p_m)
                cell_p.alignment = centered_alignment
    
   # 5. Skerspjūvio ploto F skaičiavimas naudojant Excel formules
    centered_alignment = Alignment(horizontal="center", vertical="center")
    
    idx_matmenys = None
    idx_plotas = None
    
    for idx, pavadinimas in enumerate(stulpeliai, start=1):
        if "Ortakio diametras, matmenys, m" in pavadinimas:
            idx_matmenys = idx
        if "Ortakio skerspjūvio plotas F, m2" in pavadinimas:
            idx_plotas = idx

    if idx_matmenys and idx_plotas:
        from openpyxl.utils import get_column_letter
        col_m = get_column_letter(idx_matmenys) # Matmenų stulpelio raidė
        
        if kaminas_obj.forma == 'A':
            # Apvalus: (D^2 * 3.14) / 4. D yra 6-oje eilutėje.
            # Excel formulė: =(Raidė5^2*3.14)/4
            formule_a = f"=({col_m}6^2*3.14)/4"
            cell_f = ws.cell(row=6, column=idx_plotas, value=formule_a)
            cell_f.number_format = '0.0000' # Užtikrina 4 skaičius po kablelio
            cell_f.alignment = centered_alignment
            
        else:
            # Stačiakampis: Gylis (7 eilutė) * Plotis (8 eilutė)
            # Excel formulė: =Raidė6*Raidė7
            formule_s = f"={col_m}7*{col_m}8"
            cell_f = ws.cell(row=7, column=idx_plotas, value=formule_s)
            cell_f.number_format = '0.0000'
            cell_f.alignment = centered_alignment

    # 6. Dinamiškas stulpelių A, B ir C langelių suliejimas pagal matavimo taškų kiekį
    tasku_skaicius = kaminas_obj.tasku_skaicius
    
    if tasku_skaicius > 1:
        pirma_eil = 6
        paskutine_eil = 6 + tasku_skaicius - 1

        for col in range(1, 4):  # A=1, B=2, C=3
            ws.merge_cells(
                start_row=pirma_eil,
                start_column=col,
                end_row=paskutine_eil,
                end_column=col
            )

            cell = ws.cell(row=pirma_eil, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
   
# --- NAUJA DALIS: Pito vamzdelio ir keitiklio pasirinkimas ---
    
    # 1. Duomenų struktūra (Keitiklis -> Vamzdelio Nr -> (Koeficientas K, Neapibrėžtis A))
    konfiguracija = {
        "1": { # Testo 440DP
            "7": (0.823, 0.039),
            "8": (0.827, 0.038),
            "404": (0.674, 0.036),
            "566": (0.829, 0.037)
        },
        "2": { # Testo 445
            "7": (0.825, 0.040),
            "8": (0.827, 0.079),
            "404": (0.678, 0.036),
            "566": (0.836, 0.083)
        }
    }

    print("\n--- PASIRINKITE ĮRANGĄ ---")
    keitiklis = input("Kokį slėgio keitiklį naudojote?\n1 - Testo 440DP\n2 - Testo 445 (keitiklis Nr. 0638.1445.908)\nPasirinkimas: ")
    
    if keitiklis in konfiguracija:
        vamzdelis = input("Koks Pito vamzdelio Nr., kurį naudojote matavimams (7, 8, 404 arba 566): ")
        
        if vamzdelis in konfiguracija[keitiklis]:
            koef_k, neapibreztis_a = konfiguracija[keitiklis][vamzdelis]
            
            # Surandame stulpelių indeksus
            idx_k = None
            idx_a = None
            for idx, pavadinimas in enumerate(stulpeliai, start=1):
                if "Pito vamzdelio koeficientas, K" in pavadinimas:
                    idx_k = idx
                if "Išplėstinės neapibrėžties koef. A" in pavadinimas:
                    idx_a = idx

            # Įrašome koeficientą K tiek kartų, kiek yra matavimo taškų
            if idx_k:
                for r in range(6, 6 + kaminas_obj.tasku_skaicius):
                    cell = ws.cell(row=r, column=idx_k, value=koef_k)
                    cell.number_format = '0.000'
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Įrašome neapibrėžtį A į du langelius po fraze
            if idx_a:
                for r in [6, 7]:
                    cell = ws.cell(row=r, column=idx_a, value=neapibreztis_a)
                    cell.number_format = '0.000'
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            print("Klaida: Neteisingas vamzdelio numeris.")
    else:
        print("Klaida: Neteisingas keitiklio pasirinkimas.")


    wb.save(failas)