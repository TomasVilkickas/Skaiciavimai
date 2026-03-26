import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def skaiciuoti_H2O1():
    duomenu_failas = 'Rezultatai.xlsx'
    rezultato_failas = 'H2O_rezultatai.xlsx' # Arba tas pats Rezultatai.xlsx
    
    # 1. Nuskaitome "Greitis" lapą paieškai
    df = pd.read_excel(duomenu_failas, sheet_name='Greitis', header=None)
    
    # 2. Ieškome langelio, kuriame yra "Atmosferinis slėgis"
    # Naudojame .astype(str), kad išvengtume klaidų su tuščiais langeliais
    mask = df.apply(lambda row: row.astype(str).str.contains('Atmosferinis slėgis', case=False), axis=1)
    
    found_coords = [(r, c) for r, row in enumerate(mask.values) for c, is_match in enumerate(row) if is_match]
    
    if not found_coords:
        print("Frazė 'Atmosferinis slėgis' nerasta.")
        return

    # Imame pirmą rastą atitikmenį
    row_idx, col_idx = found_coords[0]
    
    # Reikšmė yra vienu langeliu žemiau (row_idx + 1)
    reiksme = df.iloc[row_idx + 1, col_idx]
    
    # Konvertuojame į sveiką skaičių (int)
    try:
        sveikas_skaicius = int(float(reiksme))
    except (ValueError, TypeError):
        print(f"Nepavyko konvertuoti reikšmės '{reiksme}' į skaičių.")
        return

    # 3. Įrašome į Excel naudojant openpyxl (kad išlaikytume formatavimą)
    wb = load_workbook(duomenu_failas)
    
    # Patikriname, ar yra lapas "H2O", jei nėra - sukuriame
    if 'H2O' in wb.sheetnames:
        ws = wb['H2O']
    else:
        ws = wb.create_sheet('H2O')
    
    target_cell = ws['D7']
    target_cell.value = sveikas_skaicius
    
    # Centruojame langelį
    target_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Pasirenkame tą patį darbinį lapą
    ws = wb['H2O']

    # 1. Įrašome formulę į G7 langelį
    # Excel formulėse kablelis paprastai rašomas kaip taškas (0.27)
    ws['G7'] = "=+(D7-E7)*B7*C7*0.27/(273+F7)"

    # 2. Nustatome skaičių formatą (du skaičiai po kablelio)
    ws['G7'].number_format = '0.00'

    # 3. Centruojame langelį
    ws['G7'].alignment = Alignment(horizontal='center', vertical='center')

    # 1. Įrašome formulę į H7 langelį
    ws['H7'] = "=G7/1000"

    # 2. Nustatome formatą su 6 ženklais po kablelio
    # '0.000000' užtikrina, kad visada matysite šešis skaitmenis po kablelio
    ws['H7'].number_format = '0.000000'

    # 3. Centruojame langelį
    ws['H7'].alignment = Alignment(horizontal='center', vertical='center')

    # Išsaugome pakeitimus
    wb.save(duomenu_failas)
    
# Paleidžiame funkciją
if __name__ == "__main__":
    skaiciuoti_H2O()