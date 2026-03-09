import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
import numpy as np
from MatavimoVieta import Kaminas

def vykdyti_atstumai(kaminas: Kaminas):
    print(f"\n3. --- Atstumų skaičiavimo modulis ---")
    
    try:
        atstumai = []
        taskai_vienai_linijai = 0
        atstumai_raw = []

        # --- 1. DUOMENŲ NUSKAITYMAS IŠ EXCEL ---
        if kaminas.forma.upper() == 'A':
            # Apvalus ortakis
            lapas = "Apvalus 1 linija" if kaminas.liniju_skaicius == 1 else "Apvalus 2 linijos"
            df = pd.read_excel("Atstumai.xlsx", sheet_name=lapas)
            df.columns = df.columns.str.strip()
            
            col_name = 'Ortakio skersmuo, cm'
            df[col_name] = pd.to_numeric(df[col_name], errors='coerce')
            df = df.dropna(subset=[col_name]).sort_values(by=col_name)
            
            tinkamos = df[df[col_name] <= kaminas.skersmuo]
            if tinkamos.empty:
                print(f"Klaida: Skersmuo {kaminas.skersmuo} cm per mažas.")
                return
            
            eilute = tinkamos.iloc[-1]
            taskai_vienai_linijai = int(eilute['Matavimo taškų skaičius'])
            atstumai_raw = eilute.iloc[2:].values
            
        elif kaminas.forma.upper() == 'S':
            # Stačiakampis ortakis
            lapas = "Stačiakampis 1 linija" if kaminas.liniju_skaicius == 1 else "Stačiakampis 2, 3, 4, 5 linijos"
            df = pd.read_excel("Atstumai.xlsx", sheet_name=lapas)
            df.columns = df.columns.str.strip()
            
            plotas_m2 = (kaminas.gylis / 100) * (kaminas.plotis / 100)
            col_gylis = 'Ortakio gylis matavimo vietoje, cm'
            col_plotas = 'Ortakio plotas, m2'
            
            if col_plotas in df.columns:
                df[col_plotas] = df[col_plotas].astype(str).str.replace(',', '.').str.extract(r'(\d+\.?\d*)').astype(float)
            
            df[col_gylis] = pd.to_numeric(df[col_gylis], errors='coerce')
            df = df.dropna(subset=[col_gylis, col_plotas]).sort_values(by=[col_gylis, col_plotas])
            
            if plotas_m2 < 0.07:
                tinkamos = df[df[col_gylis] <= kaminas.gylis]
            else:
                tinkamos = df[(df[col_plotas] <= plotas_m2) & (df[col_gylis] <= kaminas.gylis)]
            
            if tinkamos.empty:
                print(f"Klaida: Nerasta duomenų (Plotas: {plotas_m2:.3f} m2, Gylis: {kaminas.gylis} cm).")
                return
            
            eilute = tinkamos.iloc[-1]
            taskai_vienai_linijai = int(eilute['Matavimo taškų skaičius'])
            atstumai_raw = eilute.iloc[3:].values

        # --- 2. TAŠKŲ SKAIČIAVIMO LOGIKA ---
        # Apvaliam dauginame iš linijų (ašių), stačiakampiam - ne
        if kaminas.forma.upper() == 'A':
            kaminas.tasku_skaicius = taskai_vienai_linijai * kaminas.liniju_skaicius
        else:
            kaminas.tasku_skaicius = taskai_vienai_linijai

        # --- 3. DUOMENŲ VALYMAS ---
        # Paimame tik tiek reikšmių, kiek nurodyta stulpelyje 'Matavimo taškų skaičius'
        atstumai_tikri = atstumai_raw[:taskai_vienai_linijai]

        for v in atstumai_tikri:
            try:
                v_str = str(v).replace(',', '.').strip()
                val = float(v_str)
                if not np.isnan(val):
                    atstumai.append(val)
            except (ValueError, TypeError):
                continue
        
        # Atnaujiname galutinį skaičių pagal tai, kiek realiai radome skaičių
        kaminas.tasku_skaicius = len(atstumai)

        # --- 4. ĮRAŠYMAS IR CENTRAVIMAS ---
        wb = openpyxl.load_workbook("Rezultatai.xlsx")
        ws = wb["Greitis"]
        center_style = Alignment(horizontal='center', vertical='center')
        
        # Išvalome senas reikšmes D5:D30
        for r in range(5, 31):
            cell = ws.cell(row=r, column=4)
            cell.value = None
            
        # Įrašome naujas reikšmes
        for i, verte in enumerate(atstumai):
            cell = ws.cell(row=5 + i, column=4)
            cell.value = verte
            cell.alignment = center_style
            
        wb.save("Rezultatai.xlsx")
        print(f"SĖKMINGA: Įrašyta {len(atstumai)} taškų ({kaminas.forma.upper()} forma).")

    except Exception as e:
        print(f"KRITINĖ KLAIDA: {e}")