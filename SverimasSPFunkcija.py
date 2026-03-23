import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from MatavimoVieta import Kaminas

def spalvinti_sverimas(kaminas: Kaminas):
    # Naudojame pandas ir openpyxl pagal susitarimą [cite: 2026-03-03]
    try:
        wb = openpyxl.load_workbook("Rezultatai.xlsx")
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        
    if "Svėrimas" not in wb.sheetnames:
        wb.create_sheet("Svėrimas")
    ws = wb["Svėrimas"]

    # Stiliai
    zalia_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    geltona_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    oranzine_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid") # Nauja spalva
    
    thin_side = Side(style='thin')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    centravimas = Alignment(horizontal="center", vertical="center")

    # --- NAUJAS PAPILDYMAS: Oranžinis spalvinimas ir suliejimas ---
    # Spalviname pavienius langelius
    oranziniai_langeliai = ["A6", "D6", "H6", "O6"]
    for coord in oranziniai_langeliai:
        ws[coord].fill = oranzine_fill

    # B6:B16 spalvinimas
    ws["B6"].fill = oranzine_fill

    # 1. SPALVINIMAS PAGAL FILTRŲ SKAIČIŲ
    zali_stulpeliai = [5, 6, 7, 10, 11, 12] # E, F, G, J, K, L
    
    # Filtrų dalis (nuo 6 eilutės)
    for i in range(kaminas.filtru_skaicius):
        eilute = 6 + i
        cell_c = ws.cell(row=eilute, column=3)
        cell_c.value = i + 1
        cell_c.alignment = centravimas
        
        for col in zali_stulpeliai:
            ws.cell(row=eilute, column=col).fill = zalia_fill

    # Nuspalvinama 9-ta eilutė
    for col in zali_stulpeliai:
        ws.cell(row=9, column=col).fill = zalia_fill

    # Antra žalia sekcija (nuo 11 iki 15 eilutės)
    for eilute in range(11, 16):
        for col in zali_stulpeliai:
            ws.cell(row=eilute, column=col).fill = zalia_fill

    # Geltona spalva M stulpelyje
    for eilute in list(range(6, 10)) + list(range(11, 17)):
        ws.cell(row=eilute, column=13).fill = geltona_fill

    # --- RĖMAI IR SULIEJIMAS (6-16 eilutės) ---
    s_row, e_row = 6, 16

    def apply_column_frame(col):
        for r in range(s_row, e_row + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = Border(
                left=thin_side, 
                right=thin_side, 
                top=(thin_side if r == s_row else None),
                bottom=(thin_side if r == e_row else None)
            )

    for c in [1, 2, 4, 8, 9, 14]: # Pridėti A(1) ir B(2) stulpeliai rėmams
        apply_column_frame(c)

    atskiri = [3, 5, 6, 7, 10, 11, 12, 13, 15]
    for r in range(s_row, e_row + 1):
        for c in atskiri:
            ws.cell(row=r, column=c).border = thin_border

    # --- TEKSTŲ UŽPILDYMAS IR KITA LOGIKA ---
    # (Likusi kodo dalis lieka tokia pati kaip jūsų originale)
    ws.cell(row=9, column=15).value = "Tuščiasis ėminys (mt)"
    ws.cell(row=9, column=15).alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

    papildomi_tekstai = {
        11: "Išplautos nuosėdos (mn)",
        12: "Išvalytos ėminių sistemos tuščiasis ėminys (mIt)",
        13: "Tušti indai svorio kontrolei (mIt1)",
        14: "Tušti indai svorio kontrolei (mIt2)",
        15: "Tušti indai svorio kontrolei (mIt3)",
        16: "Tuščių indų ir išvalytos ėminių sistemos tuščiojo ėminio svorio pokytis (mItp)"
    }

    for eil, tekstas in papildomi_tekstai.items():
        cell = ws.cell(row=eil, column=15)
        cell.value = tekstas
        cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")

    def gauti_galune(n, vns, dgs_kilm, dgs_vard):
        if 11 <= n % 100 <= 19: return f"{n} {dgs_kilm}"
        if n % 10 == 1: return f"{n} {vns}"
        if 2 <= n % 10 <= 9: return f"{n} {dgs_vard}"
        return f"{n} {dgs_kilm}"

    linijos_str = gauti_galune(kaminas.liniju_skaicius, "linija", "linijų", "linijos")
    filtrai_str = gauti_galune(kaminas.filtru_skaicius, "filtras", "filtrų", "filtrai")
    
    ws.cell(row=6, column=15).value = f"{linijos_str}, {filtrai_str}"
    ws.cell(row=6, column=15).alignment = Alignment(horizontal="left", vertical="center")
    ws.column_dimensions['O'].width = 38
   
    c_stulpelio_duomenys = {9: "2", 11: "41", 12: "00", 13: "01", 14: "02", 15: "03"}
    for eil, reiksme in c_stulpelio_duomenys.items():
        cell = ws.cell(row=eil, column=3)
        cell.value = reiksme
        cell.alignment = centravimas

    wb.save("Rezultatai.xlsx")