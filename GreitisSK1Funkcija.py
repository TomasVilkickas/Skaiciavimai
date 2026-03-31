import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def skaiciuoti_greitis1(kaminas_obj):
    # 1. Failo pavadinimą nustatome čia, kad nereikėtų jo siųsti iš pagrindinės programos
    failo_pavadinimas = 'Rezultatai.xlsx'
    
    try:
        # 2. Atidarome failą (naudojame openpyxl, kaip nurodyta jūsų taisyklėse)
        wb = load_workbook(failo_pavadinimas)
        
        # Patikriname, ar yra lapas „Greitis“, jei ne - naudojame pirmą aktyvų
        if 'Greitis' in wb.sheetnames:
            ws = wb['Greitis']
        else:
            ws = wb.active
            
    except FileNotFoundError:
        print(f"Klaida: Failas {failo_pavadinimas} nerastas.")
        return

    # Pagalbinė funkcija surasti frazės koordinates
    def rasti_koordinates(frazė):
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and str(frazė) in str(cell.value):
                    return cell.row, cell.column
        return None, None

    # 3. Surandame pagrindinių parametrų vietas
    atm_row, atm_col = rasti_koordinates("Atmosferinis slėgis P, hPa")
    stat_row, stat_col = rasti_koordinates("Statinis slėgis ortakyje ± ΔP, hPa")
    pk_row, pk_col = rasti_koordinates("Dujų slėgis kamine Pk= P+ ΔP, hPa")

    # Tikriname, ar viską radome
    if not all([atm_row, stat_row, pk_row]):
        print("Klaida: Nepavyko rasti visų reikiamų frazių Excel faile.")
        return

    # 4. Skaičiavimai pagal kaminas_obj parametrus
    for i in range(1, kaminas_obj.tasku_skaicius + 1):
        # Pasiimame koordinates
        src_atm_cell = ws.cell(row=atm_row + i, column=atm_col).coordinate
        src_stat_cell = ws.cell(row=stat_row + i, column=stat_col).coordinate
        
        # Įrašome formulę į Pk stulpelį
        target_cell = ws.cell(row=pk_row + i, column=pk_col)
        target_cell.value = f"={src_atm_cell}+{src_stat_cell}"
        
        # Formatuojame
        target_cell.alignment = Alignment(horizontal='center')
        target_cell.number_format = '0.0'

    # --- Dujų molinės masės Ms skaičiavimas ---
    qn_row, qn_col = rasti_koordinates("Sausų dujų tankis normaliosioms sąlygomis qn, (kg/m3)")
    ms_row, ms_col = rasti_koordinates("Dujų mol. masė Ms= qn *22.4, kg/")

    if qn_row and ms_row:
        # Paimame tikslų langelį po fraze (pvz., AA5)
        qn_langele = ws.cell(row=qn_row + 1, column=qn_col)
        qn_adresas = qn_langele.coordinate # Gauname adresą, pvz., 'AA5'
        
        for i in range(1, kaminas_obj.tasku_skaicius + 1):
            target_ms_cell = ws.cell(row=ms_row + i, column=ms_col)
            
            # SUPAPRASTINTA FORMULĖ: be papildomų vienučių kabučių, kurios dažnai sugadina failą
            target_ms_cell.value = f"={qn_adresas}*22.4"
            
            target_ms_cell.alignment = Alignment(horizontal='center')
            target_ms_cell.number_format = '0.0000'

# ---Greičio wi skaičiavimai ---

    # 1. Surandame pagrindinių parametrų koordinates (naudojame unikalias frazių dalis)
    k_row, k_col = rasti_koordinates("Pito vamzdelio koeficientas")
    t_row, t_col = rasti_koordinates("Temperatūra ortakyje t")
    atm_row, atm_col = rasti_koordinates("Atmosferinis slėgis P")
    stat_row, stat_col = rasti_koordinates("Statinis slėgis ortakyje")
    ms_row, ms_col = rasti_koordinates("Dujų mol. masė Ms")

    # K koeficientas (langelis po fraze)
    k_addr = ws.cell(row=k_row + 1, column=k_col).coordinate if k_row else "1.0"

    # 2. Ciklas per linijas (iš kaminas_obj)
    for l_idx in range(1, kaminas_obj.liniju_skaicius + 1):
        
        # Paieška konkrečiai linijai (naudojame fragmentus, kad išvengtume tarpų klaidų)
        wi_frazė = f"srauto greitis matavimo taškuose, {l_idx} linija"
        pdi_frazė = f"diferencinis slėgis taškuose" # Papildomai tikrinsime linijos nr. žemiau
        
        wi_res_row, wi_res_col = rasti_koordinates(wi_frazė)
        
        # Pdi paieška konkrečiai linijai
        pdi_row, pdi_col = None, None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "diferencinis slėgis" in str(cell.value) and f"{l_idx} linija" in str(cell.value):
                    pdi_row, pdi_col = cell.row, cell.column
                    break
            if pdi_row: break

        if wi_res_row and pdi_row:
            # 3. Einame per matavimo taškus žemyn
            for i in range(1, kaminas_obj.tasku_skaicius + 1):
                
                # Dinaminiai langelių adresai kiekvienam taškui
                pdi_addr = ws.cell(row=pdi_row + i, column=pdi_col).coordinate
                t_addr = ws.cell(row=t_row + i, column=t_col).coordinate
                atm_addr = ws.cell(row=atm_row + i, column=atm_col).coordinate
                stat_addr = ws.cell(row=stat_row + i, column=stat_col).coordinate
                ms_addr = ws.cell(row=ms_row + i, column=ms_col).coordinate
                
                target_cell = ws.cell(row=wi_res_row + i, column=wi_res_col)
                
                # Jūsų nurodyta formulė:
                # =K*129*(SQRT(Pdi*(273+t)/(Atm+Stat)/Ms))
                tiksli_formule = (
                    f"={k_addr}*129*(SQRT({pdi_addr}*(273+{t_addr})/"
                    f"({atm_addr}+{stat_addr})/{ms_addr}))"
                )
                
                target_cell.value = tiksli_formule
                
                # Formatavimas
                target_cell.alignment = Alignment(horizontal='center')
                target_cell.number_format = '0.00'
    
    # --- Vidutinio greičio w skaičiavimas ---
    # Surenkame visų užpildytų greičio langelių adresus iš visų linijų
    visi_wi_adresai = []
    
    for l_idx in range(1, kaminas_obj.liniju_skaicius + 1):
        wi_frazė = f"srauto greitis matavimo taškuose, {l_idx} linija"
        r, c = rasti_koordinates(wi_frazė)
        if r:
            for i in range(1, kaminas_obj.tasku_skaicius + 1):
                visi_wi_adresai.append(ws.cell(row=r + i, column=c).coordinate)

    # Surandame kur rašyti bendrą vidurkį
    vid_row, vid_col = rasti_koordinates("Vidutinis dujų srauto greitis ortakyje w, m/s")
    
    if vid_row and visi_wi_adresai:
        target_vid_cell = ws.cell(row=vid_row + 1, column=vid_col)
        
        # Sukonstruojame Excel formulę: =AVERAGE(A1, A2, B1, B2...)
        formule_vid = f"=AVERAGE({','.join(visi_wi_adresai)})"
        target_vid_cell.value = formule_vid
        
        # Formatavimas
        target_vid_cell.alignment = Alignment(horizontal='center')
        target_vid_cell.number_format = '0.00'
    
    # --- Debito Vk skaičiavimas ---
    # 1. Surandame vidutinio greičio w langelį (kurį ką tik užpildėme)
    w_vid_row, w_vid_col = rasti_koordinates("Vidutinis dujų srauto greitis ortakyje w, m/s")
    w_vid_addr = ws.cell(row=w_vid_row + 1, column=w_vid_col).coordinate
    # Pridedame $ ženklus fiksuotam adresui (pvz., $C$15)
    w_vid_fixed = f"${w_vid_addr[:1]}${w_vid_addr[1:]}"

    # 2. Surandame ploto F langelius (tikimės dviejų reikšmių po fraze)
    f_row, f_col = rasti_koordinates("Ortakio skerspjūvio plotas F, m2")
    
    # 3. Surandame kur rašyti rezultatus (Vk)
    vk_row, vk_col = rasti_koordinates("Dujų tūrio debitas realiomis sąlygomis Vk = wvid × F, m3/s")

    if all([w_vid_row, f_row, vk_row]):
        # Atliekame skaičiavimus dviem langeliams (i=1 ir i=2 po fraze)
        for i in range(1, 3):
            f_addr = ws.cell(row=f_row + i, column=f_col).coordinate
            target_vk_cell = ws.cell(row=vk_row + i, column=vk_col)
            
            # Įrašome formulę: f"={w_vid_fixed}*{f_addr}"
            target_vk_cell.value = f"={w_vid_fixed}*{f_addr}"
            
            # Formatavimas: centravimas ir 3 skaičiai po kablelio
            target_vk_cell.alignment = Alignment(horizontal='center')
            target_vk_cell.number_format = '0.000'
    
    # --- Debito normaliosiomis sąlygomis Vdr n.s. skaičiavimas ---
    
    # 1. Surandame bazinius parametrus (pirmosios reikšmės po frazėmis)
    p_atm_row, p_atm_col = rasti_koordinates("Atmosferinis slėgis P, hPa")
    p_stat_row, p_stat_col = rasti_koordinates("Statinis slėgis ortakyje ± ΔP, hPa")
    t_row, t_col = rasti_koordinates("Temperatūra ortakyje t, oC")
    
    # 2. Surandame kintamąjį Vk (du langeliai, kuriuos ką tik užpildėme)
    vk_res_row, vk_res_col = rasti_koordinates("Dujų tūrio debitas realiomis sąlygomis Vk = wvid × F, m3/s")
    
    # 3. Surandame kur rašyti galutinį rezultatą Vdr n.s.
    vdr_row, vdr_col = rasti_koordinates("Dujų tūrio debitas normaliosiomis sąlygomis Vdr n.s =Vk *0,269*(P±ΔP)/(273+t), m3/s")

    if all([p_atm_row, p_stat_row, t_row, vk_res_row, vdr_row]):
        # Paruošiame fiksuotus adresus (naudojame pirmuosius langelius po frazėmis)
        def fix_addr(r, c):
            addr = ws.cell(row=r + 1, column=c).coordinate
            return f"${addr[:1]}${addr[1:]}"

        p_fixed = fix_addr(p_atm_row, p_atm_col)
        stat_fixed = fix_addr(p_stat_row, p_stat_col)
        t_fixed = fix_addr(t_row, t_col)

        # Skaičiuojame abiems Vk reikšmėms
        for i in range(1, 3):
            vk_addr = ws.cell(row=vk_res_row + i, column=vk_res_col).coordinate
            target_vdr_cell = ws.cell(row=vdr_row + i, column=vdr_col)
            
            # Formulė: Vk * 0.269 * (P + ΔP) / (273 + t)
            tiksli_vdr_formule = f"={vk_addr}*0.269*({p_fixed}+{stat_fixed})/(273+{t_fixed})"
            
            target_vdr_cell.value = tiksli_vdr_formule
            
            # Formatavimas
            target_vdr_cell.alignment = Alignment(horizontal='center')
            target_vdr_cell.number_format = '0.000'
    
    # --- Vandens garų koncentracijos x skaičiavimas ---
    
    # 1. Surandame reikiamų parametrų koordinates
    m_h2o_row, m_h2o_col = rasti_koordinates("Vandens kondensato masė m H2O, kg")
    vmn_row, vmn_col = rasti_koordinates("Prasiurbtas dujų tūris normaliosiomis sąlygomis Vmn, Nm3")
    
    # 2. Surandame kur rašyti rezultatą x
    x_row, x_col = rasti_koordinates("Vandens garų konc. dujose x=mH2O/Vmn*qn, kg/kg")

    if all([m_h2o_row, vmn_row, x_row]):
        # Atliekame skaičiavimus dviem eilutėms (i=1 ir i=2)
        for i in range(1, 3):
            m_addr = ws.cell(row=m_h2o_row + i, column=m_h2o_col).coordinate
            vmn_addr = ws.cell(row=vmn_row + i, column=vmn_col).coordinate
            target_x_cell = ws.cell(row=x_row + i, column=x_col)
            
            # Formulė: mH2O / (1.29 * Vmn)
            # Naudojame tašką kaip skirtuką skaičiui 1.29
            tiksli_x_formule = f"={m_addr}/(1.29*{vmn_addr})"
            
            target_x_cell.value = tiksli_x_formule
            
            # Formatavimas: centravimas ir 4 skaičiai po kablelio (dažniausiai x reikalauja didesnio tikslumo)
            target_x_cell.alignment = Alignment(horizontal='center')
            target_x_cell.number_format = '0.0000'
    
    # --- Sausų dujų tūrio debito Vn skaičiavimas ---
    
    # 1. Surandame pradinių duomenų koordinates
    vdr_ns_row, vdr_ns_col = rasti_koordinates("Dujų tūrio debitas normaliosiomis sąlygomis Vdr n.s =Vk *0,269*(P±ΔP)/(273+t), m3/s")
    x_konc_row, x_konc_col = rasti_koordinates("Vandens garų konc. dujose x=mH2O/Vmn*qn, kg/kg")
    
    # 2. Surandame kur rašyti rezultatą Vn
    vn_row, vn_col = rasti_koordinates("Sausų dujų tūrio debitas normaliosiomis sąlygomis Vn= Vdr n.s *((1/(1+(x*qn/0.8038)))")

    if all([vdr_ns_row, x_konc_row, vn_row]):
        # Atliekame skaičiavimus dviem eilutėms (i=1 ir i=2)
        for i in range(1, 3):
            vdr_addr = ws.cell(row=vdr_ns_row + i, column=vdr_ns_col).coordinate
            x_addr = ws.cell(row=x_konc_row + i, column=x_konc_col).coordinate
            target_vn_cell = ws.cell(row=vn_row + i, column=vn_col)
            
            # Formulė: Vdr_ns * (1 / (1 + (x * 1.29 / 0.8038)))
            # Naudojame taškus skaičių skyrikliams
            tiksli_vn_formule = f"={vdr_addr}*(1/(1+({x_addr}*1.29/0.8038)))"
            
            target_vn_cell.value = tiksli_vn_formule
            
            # Formatavimas: centravimas ir 3 skaičiai po kablelio
            target_vn_cell.alignment = Alignment(horizontal='center')
            target_vn_cell.number_format = '0.000'

    # --- Išplėstinės neapibrėžties Uv skaičiavimas ---
    
    # 1. Surandame pradinių duomenų koordinates
    vn_res_row, vn_res_col = rasti_koordinates("Sausų dujų tūrio debitas normaliosiomis sąlygomis Vn= Vdr n.s *((1/(1+(x*qn/0.8038)))")
    koef_a_row, koef_a_col = rasti_koordinates("Išplėstinės neapibrėžties koef. A")
    
    # 2. Surandame kur rašyti rezultatą Uv
    uv_row, uv_col = rasti_koordinates("Išplėstinė neapibrėžtis (dujų tūrio debitas) Uv")

    if all([vn_res_row, koef_a_row, uv_row]):
        # Paruošiame fiksuotą koeficiento A adresą (pirmas langelis po fraze)
        a_addr = ws.cell(row=koef_a_row + 1, column=koef_a_col).coordinate
        a_fixed = f"${a_addr[:1]}${a_addr[1:]}"

        # Atliekame skaičiavimus dviem eilutėms (i=1 ir i=2)
        for i in range(1, 3):
            vn_addr = ws.cell(row=vn_res_row + i, column=vn_res_col).coordinate
            target_uv_cell = ws.cell(row=uv_row + i, column=uv_col)
            
            # Formulė: Vn * A
            target_uv_cell.value = f"={vn_addr}*{a_fixed}"
            
            # Formatavimas: centravimas ir 3 skaičiai po kablelio
            target_uv_cell.alignment = Alignment(horizontal='center')
            target_uv_cell.number_format = '0.000'
    
   # --- Išplėstinės neapibrėžties Uw skaičiavimas (PATAISYTA) ---
    
    # 1. Surandame koordinates
    w_vid_row, w_vid_col = rasti_koordinates("Vidutinis dujų srauto greitis ortakyje w, m/s")
    koef_a_row, koef_a_col = rasti_koordinates("Išplėstinės neapibrėžties koef. A")
    uw_row, uw_col = rasti_koordinates("Išplėstinė neapibrėžtis (dujų srauto greitis) Uw")

    if all([w_vid_row, koef_a_row, uw_row]):
        # Užfiksuojame TIK vidutinį greitį w (nes jis vienas)
        w_addr = ws.cell(row=w_vid_row + 1, column=w_vid_col).coordinate
        w_fixed = f"${w_addr[:1]}${w_addr[1:]}"
        
        # Ciklas per dvi eilutes neapibrėžčiai A ir rezultatui Uw
        for i in range(1, 3):
            # Šis adresas keisis (A1, A2...), todėl nenaudojame $
            a_addr = ws.cell(row=koef_a_row + i, column=koef_a_col).coordinate
            target_uw_cell = ws.cell(row=uw_row + i, column=uw_col)
            
            # Formulė: w_fixed (konstanta) * a_addr (kintantis)
            target_uw_cell.value = f"={w_fixed}*{a_addr}"
            
            # Formatavimas
            target_uw_cell.alignment = Alignment(horizontal='center')
            target_uw_cell.number_format = '0.00'
    
    # 5. Išsaugome
    wb.save(failo_pavadinimas)