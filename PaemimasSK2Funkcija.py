from openpyxl import load_workbook
from openpyxl.styles import Alignment
from MatavimoVieta import Kaminas

def skaiciuoti_paemimas2(kaminas_obj: Kaminas):
    failas = "Rezultatai.xlsx"
    wb = load_workbook(failas)
    
    ws_greitis = wb["Greitis"]
    ws_paemimas = wb["Paėmimas"]
    
    # 1. Randame stulpelius "Greitis" lape (ieškome 5-oje eilutėje)
    antraštės_eilutė = 5
    linijų_stulpeliai = {} # Pvz: {1: 7, 2: 15} (linijos numeris: stulpelio indeksas)
    
    # Pereiname per stulpelius (pvz., iki 50), kad rastume tuos "wi ="
    for col in range(1, 51):
        cell_value = str(ws_greitis.cell(row=antraštės_eilutė, column=col).value or "")
        for l in range(1, kaminas_obj.liniju_skaicius + 1):
            if f"{l} linija" in cell_value and "wi =" in cell_value:
                linijų_stulpeliai[l] = col

    # 2. Perkeliame nuorodas į "Paėmimas" lapą
    dabartine_eilute_paemimas = 6
    stulpelis_g_paemimas = 7 # G stulpelis
    
    # Duomenys "Greitis" lape prasideda nuo 6 eilutės
    pradine_duomenu_eilute_greitis = 6

    for l in range(1, kaminas_obj.liniju_skaicius + 1):
        if l in linijų_stulpeliai:
            source_col_idx = linijų_stulpeliai[l]
            # Konvertuojame stulpelio skaičių į Excel raidę (pvz., 7 -> G)
            col_letter = ws_greitis.cell(row=antraštės_eilutė, column=source_col_idx).column_letter
            
            for i in range(kaminas_obj.tasku_skaicius):
                row_in_greitis = pradine_duomenu_eilute_greitis + i
                target_cell = ws_paemimas.cell(row=dabartine_eilute_paemimas, column=stulpelis_g_paemimas)
                
                # Sukuriame formulę-nuorodą: ='Greitis'!G6
                target_cell.value = f"='Greitis'!{col_letter}{row_in_greitis}"
                
                # Formatavimas
                target_cell.alignment = Alignment(horizontal='center', vertical='center')
                target_cell.number_format = '0.00'
                
                dabartine_eilute_paemimas += 1
            
            # Praleidžiame 4 tuščias eilutes po kiekvienos linijos bloko
            dabartine_eilute_paemimas += 4
            
    wb.save(failas)