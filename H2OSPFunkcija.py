import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

def spalvinti_H2O():
    
    failo_pavadinimas = "Rezultatai.xlsx"
    wb = openpyxl.load_workbook(failo_pavadinimas)
    
    if "H2O" not in wb.sheetnames:
        wb.create_sheet("H2O")
    ws = wb["H2O"]

    # 1. Apibrėžiame stilius
    plonas_remelis = Side(border_style="thin", color="000000")
    border = Border(left=plonas_remelis, right=plonas_remelis, top=plonas_remelis, bottom=plonas_remelis)
    
    bold_font = Font(bold=True)
    centravimas = Alignment(horizontal="center", vertical="center")
    
    # Spalvos
    zalia_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    oranzine_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    geltona_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 2. Rėminame 6, 8, 9, 10 eilutes (A-L stulpeliai)
    reminamos_eilutes = [6, 8, 9, 10]
    for eilute in reminamos_eilutes:
        for stulpelis in range(1, 13):  # 1 iki 12 (A iki L)
            cell = ws.cell(row=eilute, column=stulpelis)
            cell.border = border

    # 3. 7-os eilutės formatavimas (A-L)
    stulpeliu_nustatymai = {
        1:  {"fill": None,          "format": "General",   "text": "TŠ 001 H2O"},# A
        2:  {"fill": zalia_fill,    "format": "0"},                              # B
        3:  {"fill": zalia_fill,    "format": "0"},                              # C
        4:  {"fill": oranzine_fill, "format": "0"},                              # D
        5:  {"fill": zalia_fill,    "format": "0.0"},                            # E
        6:  {"fill": zalia_fill,    "format": "0.0"},                            # F
        7:  {"fill": geltona_fill,  "format": "0.00"},                           # G
        8:  {"fill": geltona_fill,  "format": "0.000000"},                       # H
        9:  {"fill": zalia_fill,    "format": "0.000"},                          # I
        10: {"fill": oranzine_fill, "format": "0.00"},                           # J
        11: {"fill": geltona_fill,  "format": "0.000"},                          # K
        12: {"fill": None,          "format": "General"}                         # L
    }

    for col_idx, settings in stulpeliu_nustatymai.items():
        cell = ws.cell(row=7, column=col_idx)
        cell.border = border
        cell.font = bold_font
        cell.alignment = centravimas
        
        # Įrašome tekstą A stulpelyje
        if "text" in settings:
            cell.value = settings["text"]
        
        # Nustatome užpildą
        if settings["fill"]:
            cell.fill = settings["fill"]
            
        # Nustatome skaičių formatą
        cell.number_format = settings["format"]

    wb.save(failo_pavadinimas)