import openpyxl
from openpyxl.styles import Alignment
from MatavimoVieta import Kaminas

def patikrinti_atminti():
    print(f"\n4. --- Klasės atminties tikrinimas ---")
    failas = "Rezultatai.xlsx"
    try:
        wb = openpyxl.load_workbook(failas)
        ws = wb["Greitis"]
        center = Alignment(horizontal='center', vertical='center')
        
        duomenys = [
            ("PARAMETRAS", "REIKŠMĖ"),
            ("Forma", Kaminas.forma),
            ("Gylis (cm)", Kaminas.gylis),
            ("Plotis (cm)", getattr(Kaminas, 'plotis', '-')),
            ("Linijų skaičius", Kaminas.liniju_skaicius),
            ("Filtrų skaičius", Kaminas.filtru_skaicius),
            ("Iš viso taškų", getattr(Kaminas, 'tasku_skaicius', 0))
        ]
        
        for i, (p, r) in enumerate(duomenys, start=1):
            ws.cell(row=i, column=1, value=p).alignment = center
            ws.cell(row=i, column=2, value=r).alignment = center
            
        wb.save(failas)
        wb.close() # Uždarome failą
        print("A ir B stulpeliai užpildyti.")
    except Exception as e:
        print(f"Klaida: {e}")