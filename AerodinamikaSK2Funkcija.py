import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from MatavimoVieta import Kaminas

def skaiciuoti_aerodinamika2(kaminas_obj: Kaminas):
    failo_pavadinimas = 'Rezultatai.xlsx'
    lapas_saltinis = 'Paėmimas'
    lapas_tikslas = 'Aerodinamika'
    
    try:
        # Naudojame openpyxl darbui su formulėmis [cite: 2026-03-03]
        wb = load_workbook(failo_pavadinimas, data_only=False)
    except FileNotFoundError:
        print(f"Klaida: Failas {failo_pavadinimas} nerastas.")
        return

    ws_tikslas = wb[lapas_tikslas]
    
    t_skaicius = kaminas_obj.tasku_skaicius
    l_skaicius = kaminas_obj.liniju_skaicius
    f_skaicius = kaminas_obj.filtru_skaicius
    viso_lenteliu = l_skaicius * f_skaicius
    
    tikslo_eilute = 6
    saltinio_eilute = 6 
    visos_sumos_adresai = []

    for lenteles_idx in range(viso_lenteliu):
        # Išsisaugome pradžios eilutę sumos diapazonui
        pradzios_eilute = tikslo_eilute
        
        for i in range(t_skaicius):
            # --- I STULPELIS ---
            formule_i = (
                f"=(E{tikslo_eilute}/(1-(H{tikslo_eilute}/1000))*((273+'{lapas_saltinis}'!F{saltinio_eilute})/273)*"
                f"(1013/('{lapas_saltinis}'!H{saltinio_eilute}+'{lapas_saltinis}'!I{saltinio_eilute})))"
            )
            cell_i = ws_tikslas.cell(row=tikslo_eilute, column=9)
            cell_i.value = formule_i
            cell_i.alignment = Alignment(horizontal='center', vertical='center')
            cell_i.number_format = '0.000'

            # --- J STULPELIS ---
            formule_j = (
                f"=(I{tikslo_eilute}/('{lapas_saltinis}'!M{saltinio_eilute}*60)/"
                f"((3.142*(('{lapas_saltinis}'!K{saltinio_eilute}/1000)*"
                f"('{lapas_saltinis}'!K{saltinio_eilute}/1000))/4)))"
            )
            cell_j = ws_tikslas.cell(row=tikslo_eilute, column=10)
            cell_j.value = formule_j
            cell_j.alignment = Alignment(horizontal='center', vertical='center')
            cell_j.number_format = '0.000'
            
            tikslo_eilute += 1
            saltinio_eilute += 1
        
        # --- SUMOS SKAIČIAVIMAS (I stulpelyje, iškart po lentelės) ---
        pabaigos_eilute = tikslo_eilute - 1
        sumos_cell = ws_tikslas.cell(row=tikslo_eilute, column=9)
        
        # Sukuriame Excel SUM formulę: =SUM(I{nuo}:I{iki})
        sumos_cell.value = f"=SUM(I{pradzios_eilute}:I{pabaigos_eilute})"

        visos_sumos_adresai.append(sumos_cell.coordinate)
        
        # Formatavimas: Centruota, 3 ženklai po kablelio, paryškinta
        sumos_cell.alignment = Alignment(horizontal='center', vertical='center')
        sumos_cell.number_format = '0.000'
        sumos_cell.font = Font(bold=True) # Pridėjau paryškinimą, kad suma išsiskirtų
        
        # Po lentelės ir sumos eilutės darome likusį tarpą (4 - 1 sumos eilutė = 3 eilutės)
        tikslo_eilute += 4
        saltinio_eilute += 4

    # Sumų sumos skaičiavimai
    # Praleidžiame 3 eilutes po paskutinės lentelės sumos
    # (Kadangi po ciklo tikslo_eilute jau yra paslinkta +4, atimame 1, kad gautume tiksliai 3 tarpus)
    galutine_eilute = tikslo_eilute - 1
    galutine_suma_cell = ws_tikslas.cell(row=galutine_eilute, column=9)
    
    # Sujungiame visų sumų adresus į vieną formulę: =I12+I25+I38...
    galutine_suma_cell.value = f"={'+'.join(visos_sumos_adresai)}"
    
    # Formatavimas galutiniam rezultatui
    galutine_suma_cell.alignment = Alignment(horizontal='center', vertical='center')
    galutine_suma_cell.number_format = '0.000'
    galutine_suma_cell.font = Font(bold=True, color="FF0000") # Raudona spalva akcentui

    wb.save(failo_pavadinimas)