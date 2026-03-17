import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

def nuskaityti_ir_perkelti_Paemimas(kaminas_obj):
    failas = "Pradiniai.xlsx"
    wb = load_workbook(failas)
    ws = wb["Paėmimas"] if "Paėmimas" in wb.sheetnames else wb.create_sheet("Paėmimas")

    laukai = {
        "A": {"pavadinimas": "Išmatuota O2 koncentracija (%)", "formatas": "0.00", "plotis": 15},
        "B": {"pavadinimas": "Išmatuota CO2 koncentracija (%)", "formatas": "0.00", "plotis": 15},
        "C": {"pavadinimas": "Temperatūra ortakyje tor (°C)", "formatas": "0.0", "plotis": 14}
    }

    thin_side = Side(border_style="thin")
    remelis = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    centravimas = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_letter, nustatymai in laukai.items():
        cell_header = ws[f"{col_letter}2"]
        cell_header.value = nustatymai["pavadinimas"]
        cell_header.font = Font(bold=True)
        cell_header.alignment = centravimas
        cell_header.border = remelis
        ws.column_dimensions[col_letter].width = nustatymai["plotis"]

        for row in range(3, 6):
            cell_data = ws.cell(row=row, column=cell_header.column)
            cell_data.border = remelis
            cell_data.alignment = centravimas
            cell_data.number_format = nustatymai["formatas"]

    wb.save(failas)

def perkelti_paemimas_duomenis(kaminas_obj=None):
    failas_prad = "Pradiniai.xlsx"
    failas_rez = "Rezultatai.xlsx"
    
    wb_prad = load_workbook(failas_prad, data_only=True)
    ws_prad = wb_prad["Paėmimas"]
    
    wb_rez = load_workbook(failas_rez)
    ws_rez = wb_rez["Paėmimas"]

    stulpeliu_map = {}
    temp_count = 0  # Skaitiklis temperatūros stulpeliams
    
    # 1. PAIEŠKA (O2, CO2 ir antra Temperatūra)
    for cell in ws_rez[5]:
        if cell.value is None: continue
        txt = str(cell.value).lower().replace("₂", "2").strip()
        
        if "išmatuota" in txt:
            if "co2" in txt:
                stulpeliu_map["CO2"] = cell.column
            elif "o2" in txt:
                stulpeliu_map["O2"] = cell.column
           
              # ŠITA DALIS PAKEISTA:
        # Temperatūros ieškome nepriklausomai nuo to, ar yra žodis "išmatuota".
        # Taip pat pridėtas skaitiklis (temp_count), kad praleistų pirmąjį stulpelį.
        if "temperatūra" in txt or "tor" in txt:
            temp_count += 1
            if temp_count == 3: # <--- Čia užtikriname, kad imtų tik TREČIĄ sutaptį
                stulpeliu_map["TEMP"] = cell.column

    # 2. DUOMENŲ PERKĖLIMAS (Skaitymas iš A, B, C | Rašymas į stulpelius pagal map)
    darbo_laukai = {"A": "O2", "B": "CO2", "C": "TEMP"}
    
    thin_side = Side(border_style="thin")
    remelis = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    centravimas = Alignment(horizontal='center', vertical='center')

    for prad_col, raktas in darbo_laukai.items():
        if raktas in stulpeliu_map:
            rez_col_index = stulpeliu_map[raktas]
            
            # row_prad eina per 3, 4, 5 (Pradiniai.xlsx)
            for row_prad in range(3, 6):
                # row_rez paskaičiuojamas pridedant 3 (kad gautume 6, 7, 8)
                row_rez = row_prad + 3 
                
                val = ws_prad[f"{prad_col}{row_prad}"].value
                target = ws_rez.cell(row=row_rez, column=rez_col_index)
                
                target.value = val
                target.alignment = centravimas
                target.border = remelis
                # Temperatūrai vienas skaičius po kablelio, kitiems du
                target.number_format = '0.0' if raktas == "TEMP" else '0.00'
            
            # Vidurkio formulė (lieka 9 eilutėje)
            col_let = get_column_letter(rez_col_index)
            v_cell = ws_rez.cell(row=9, column=rez_col_index)
            v_cell.value = f"=AVERAGE({col_let}6:{col_let}8)"
            v_cell.font = Font(bold=False)
            v_cell.alignment = centravimas
            v_cell.border = remelis
            v_cell.number_format = target.number_format
    
    wb_rez.save(failas_rez)