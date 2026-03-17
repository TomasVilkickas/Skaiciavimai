import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

def spalvinti_greitis(kaminas):
    file_name = "Rezultatai.xlsx"
    sheet_name = "Greitis"
    
    try:
        wb = openpyxl.load_workbook(file_name)
        ws = wb[sheet_name]
    except Exception as e:
        print(f"Klaida atidarant failą: {e}")
        return

    # --- NUSTATYMAI ---
    zalia_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    geltona_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    oranzine_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    thin_side = Side(style="thin")
    pilnas_remelis = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    start_row = 6
    tasku_skaicius = getattr(kaminas, 'tasku_skaicius', 0)
    end_row = start_row + tasku_skaicius - 1
    
    # Tikrinimo frazės
    tik_zalia = ["diferencinis slėgis"]
    tik_oranzine = ["temperatūra ortakyje"]
    miksas_zalia_oranzine = ["pito vamzdelio koeficientas", "atmosferinis slėgis", "statinis slėgis ortakyje"]
    tik_geltona = ["dujų slėgis kamine", "dujų mol. masė", "dujų srauto greitis matavimo taškuose"]

    # --- 1. KINTAMOSIOS DALIES SPALVINIMAS IR TIKSLUS GALO NUSTATYMAS ---
    # Naudojame šį kintamąjį kaip „inkarą“ tolimesniems stulpeliams
    paskutinis_duomenu_col = 4 

    # Einame per stulpelius nuo D (4) iki lapo galo
    for col in range(4, ws.max_column + 1):
        header_val = str(ws.cell(row=5, column=col).value or "").lower().strip()
        
        # Tikriname, ar šis stulpelis yra kintamosios dalies dalis
        yra_kintamas = any(f in header_val for f in tik_zalia + tik_oranzine + miksas_zalia_oranzine + tik_geltona)
        
        if yra_kintamas:
            paskutinis_duomenu_col = col # Fiksuojame vėliausią rastą duomenų stulpelį
            
            for r in range(start_row, end_row + 1):
                target_cell = ws.cell(row=r, column=col)
                target_cell.border = pilnas_remelis
                
                if any(f in header_val for f in tik_zalia):
                    target_cell.fill = zalia_fill
                elif any(f in header_val for f in tik_oranzine):
                    target_cell.fill = oranzine_fill
                elif any(f in header_val for f in miksas_zalia_oranzine):
                    target_cell.fill = zalia_fill if r == start_row else oranzine_fill
                elif any(f in header_val for f in tik_geltona):
                    target_cell.fill = geltona_fill

    # --- 2. STACIONARŪS STULPELIAI (Prasideda iškart po paskutinio kintamo) ---
    start_col = paskutinis_duomenu_col + 1
    
    # 5 stulpelių blokas (5x4)
    konfig = {
        0: ([1, 0, 0, 0], geltona_fill),
        1: ([1, 1, 1, 0], zalia_fill),
        2: ([1, 1, 0, 0], geltona_fill),
        3: ([1, 1, 0, 0], geltona_fill),
        4: ([1, 1, 0, 0], geltona_fill)
    }

    for c_off in range(5):
        curr_c = start_col + c_off
        planas, spalva = konfig[c_off]
        for r_off in range(4):
            cell = ws.cell(row=6 + r_off, column=curr_c)
            cell.border = pilnas_remelis
            cell.fill = spalva if planas[r_off] else PatternFill(fill_type=None)

    # 7 stulpelių blokas (7x2)
    extra_col = start_col + 5
    sekos_spalvos = [oranzine_fill, geltona_fill, geltona_fill, oranzine_fill, zalia_fill, geltona_fill, geltona_fill]
    for i, spalva in enumerate(sekos_spalvos):
        curr_c = extra_col + i
        for r_off in range(2):
            cell = ws.cell(row=6 + r_off, column=curr_c)
            cell.fill = spalva
            cell.border = pilnas_remelis

    # --- 3. SPECIFINĖ PABAIGA ---
    # Oranžinis langelis
    fin_col = extra_col + 7
    ws.cell(row=6, column=fin_col).fill = oranzine_fill
    ws.cell(row=6, column=fin_col).border = pilnas_remelis
    
    # Sulietas tuščias langelis (6-7 eilutės)
    merge_col = fin_col + 1
    ws.merge_cells(start_row=6, start_column=merge_col, end_row=7, end_column=merge_col)
    for r in [6, 7]:
        cell = ws.cell(row=r, column=merge_col)
        cell.border = pilnas_remelis
        cell.fill = PatternFill(fill_type=None)

    # --- 4. APATINĖS EILUTĖS (Naudojant tą patį paskutinis_duomenu_col) ---
    # Rėmeliai pirmoje eilutėje po lentelės iki kintamos dalies galo
    eilute_po = end_row + 1
    for col in range(1, paskutinis_duomenu_col + 1):
        ws.cell(row=eilute_po, column=col).border = pilnas_remelis

    # Oranžinė RPU eilutė
    rpu_row = eilute_po + 1
    # A, B, C rėmai
    for col in range(1, 4):
        ws.cell(row=rpu_row, column=col).border = pilnas_remelis
    
    # RPU langelis (D stulpelis)
    ws.cell(row=rpu_row, column=4).value = "RPU***"
    ws.cell(row=rpu_row, column=4).border = pilnas_remelis
    
    # Oranžinis tekstas
    if paskutinis_duomenu_col >= 5:
        ws.merge_cells(start_row=rpu_row, start_column=5, end_row=rpu_row, end_column=paskutinis_duomenu_col)
        o_cell = ws.cell(row=rpu_row, column=5)
        o_cell.value = "Kai bus 2 linijos, RPU*** yra 2 linijos vidurinis taškas"
        o_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        o_cell.alignment = Alignment(horizontal="center")
        for col in range(5, paskutinis_duomenu_col + 1):
            ws.cell(row=rpu_row, column=col).border = pilnas_remelis

            # --- FIX: A-D stulpelių rėmeliai visai lentelei ---
        for col in range(1, 5):  # A=1, B=2, C=3, D=4
            for r in range(start_row, end_row + 1):
                cell = ws.cell(row=r, column=col)
                cell.border = pilnas_remelis
    
                wb.save(file_name)