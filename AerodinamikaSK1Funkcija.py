import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from MatavimoVieta import Kaminas

def skaiciuoti_aerodinamika1(kaminas_obj):
    failo_pavadinimas = 'Rezultatai.xlsx'
    lapas_saltinis = 'Paėmimas'
    lapas_tikslas = 'Aerodinamika'
    
    # Atidarome failą (visada naudokite pandas ir openpyxl)
    try:
        wb = load_workbook(failo_pavadinimas, data_only=False)
    except FileNotFoundError:
        print(f"Klaida: Failas {failo_pavadinimas} nerastas.")
        return

    if lapas_saltinis not in wb.sheetnames or lapas_tikslas not in wb.sheetnames:
        return

    ws_tikslas = wb[lapas_tikslas]
    
    # Parametrai iš kaminas_obj
    t_skaicius = kaminas_obj.tasku_skaicius
    l_skaicius = kaminas_obj.liniju_skaicius
    f_skaicius = kaminas_obj.filtru_skaicius
    viso_lenteliu = l_skaicius * f_skaicius
    
    tikslo_eilute = 6
    saltinio_eilute = 6 
    vidurkiu_langeliai = []
    sumu_langeliai = []

    # --- Pastabose esančių linijų ir filtrų numeracija ---
    def lietuvinti(n, vns, dgs):
        if n % 10 == 1 and n % 100 != 11:
            return vns
        return dgs

    pastabu_eilute = 6  # Pradedame nuo N6
    # ----------------------------------------------

    for lenteles_idx in range(viso_lenteliu):
        # Pataisyta logika: linijos kinta pirmiau
        filtro_nr = (lenteles_idx // l_skaicius) + 1
        linijos_nr = (lenteles_idx % l_skaicius) + 1
        
        # Sudarome tekstą
        l_tekstas = lietuvinti(linijos_nr, "linija", "linijos")
        f_tekstas = lietuvinti(filtro_nr, "filtras", "filtrai")
        irasas = f"{linijos_nr} {l_tekstas}, {filtro_nr} {f_tekstas}"
        
        # Įrašome į N stulpelį (14-as stulpelis)
        cell_n = ws_tikslas.cell(row=pastabu_eilute, column=14)
        cell_n.value = irasas
        cell_n.alignment = Alignment(horizontal='left', vertical='center')
        
        # Atnaujiname eilutę kitam įrašui
        pastabu_eilute += (t_skaicius + 4)

    for lenteles_idx in range(viso_lenteliu):
        blokas_nuo = tikslo_eilute
        blokas_iki = tikslo_eilute + t_skaicius - 1

        for i in range(t_skaicius):
            # --- D STULPELIS (3 sk.) ---
            cell_d = ws_tikslas.cell(row=tikslo_eilute, column=4)
            cell_d.value = f"='{lapas_saltinis}'!L{saltinio_eilute}"
            cell_d.alignment = Alignment(horizontal='center')
            cell_d.number_format = '0.000'
            
            # --- E STULPELIS (6 sk.) ---
            formule_e = (f"=D{tikslo_eilute}*(273/(273+'{lapas_saltinis}'!O{saltinio_eilute})*"
                         f"('{lapas_saltinis}'!H{saltinio_eilute}-'{lapas_saltinis}'!J{saltinio_eilute})/1013)")
            cell_e = ws_tikslas.cell(row=tikslo_eilute, column=5)
            cell_e.value = formule_e
            cell_e.alignment = Alignment(horizontal='center')
            cell_e.number_format = '0.000000'
            
            # --- F STULPELIS (Suma) ---
            if i == 0:
                cell_f = ws_tikslas.cell(row=tikslo_eilute, column=6)
                cell_f.value = f"=SUM(E{blokas_nuo}:E{blokas_iki})"
                cell_f.alignment = Alignment(horizontal='center')
                cell_f.number_format = '0.000000'
                # Įsimename šį langelį suvestinei
                sumu_langeliai.append(f"F{tikslo_eilute}")
            
            # --- G STULPELIS (2 sk. po kablelio) ---
            # Formulė: ((Q/100)*1.965)+((P/100)*1.429)+((1-(Q/100)-(P/100))*1.251)
            formule_g = (f"=(('{lapas_saltinis}'!Q{saltinio_eilute}/100)*1.965)+"
                         f"(('{lapas_saltinis}'!P{saltinio_eilute}/100)*1.429)+"
                         f"((1-('{lapas_saltinis}'!Q{saltinio_eilute}/100)-('{lapas_saltinis}'!P{saltinio_eilute}/100))*1.251)")
            cell_g = ws_tikslas.cell(row=tikslo_eilute, column=7) # 7 = G
            cell_g.value = formule_g
            cell_g.alignment = Alignment(horizontal='center')
            cell_g.number_format = '0.00'
            
            tikslo_eilute += 1
            saltinio_eilute += 1
        
        #Vidurkio skaičiavimas    
        cell_g_vidurkis = ws_tikslas.cell(row=tikslo_eilute, column=7)
        cell_g_vidurkis.value = f"=AVERAGE(G{blokas_nuo}:G{blokas_iki})"
        cell_g_vidurkis.alignment = Alignment(horizontal='center')
        cell_g_vidurkis.number_format = '0.00'
        vidurkiu_langeliai.append(f"G{tikslo_eilute}")
        
        tikslo_eilute += 4
        saltinio_eilute += 4

        # --- SUVESTINĖ: Vidurkių vidurkis G stulpelyje ---
    # tikslo_eilute po ciklo pabaigos jau yra per 4 eilutes žemiau paskutinės lentelės vidurkio.
    # SP1 funkcija naudoja 'dabartine_eilute -= 4' ir tada '+ 3', todėl formulę rašome čia:
    suvestines_g_eilute = tikslo_eilute - 1 

    if vidurkiu_langeliai:
        # Suformuojame AVERAGE formulę iš visų surinktų individualių vidurkių
        vidurkiu_sarasas = ",".join(vidurkiu_langeliai)
        cell_suvestine_g = ws_tikslas.cell(row=suvestines_g_eilute, column=7)
        cell_suvestine_g.value = f"=AVERAGE({vidurkiu_sarasas})"
        cell_suvestine_g.alignment = Alignment(horizontal='center', vertical='center')
        cell_suvestine_g.number_format = '0.00'

    # --- SUVESTINĖ: Viso suma F stulpelyje ---
    # Šis langelis pagal SP1 logiką yra iškart po paskutinės lentelės blokų
    suvestines_f_eilute = tikslo_eilute - 4 

    if sumu_langeliai:
        # Suformuojame SUM formulę iš visų lentelių pirmųjų eilučių (F stulpelis)
        sumu_sarasas = ",".join(sumu_langeliai)
        cell_suvestine_f = ws_tikslas.cell(row=suvestines_f_eilute, column=6)
        cell_suvestine_f.value = f"=SUM({sumu_sarasas})"
        cell_suvestine_f.alignment = Alignment(horizontal='center', vertical='center')
        cell_suvestine_f.number_format = '0.000000'

    # --- Vidurkių vidurkio perkėlimas į H2O lapą ---
    if 'H2O' in wb.sheetnames:
        ws_h2o = wb['H2O']
        cell_j7 = ws_h2o['J7']
        
        # Įrašome nuorodą į Aerodinamika lapo suvestinės langelį
        # suvestines_g_eilute yra ta pati eilutė, kurioje ką tik įrašėme vidurkį
        cell_j7.value = f"=Aerodinamika!G{suvestines_g_eilute}"
        
        # Formatavimas: centravimas ir 2 skaičiai po kablelio
        cell_j7.alignment = Alignment(horizontal='center', vertical='center')
        cell_j7.number_format = '0.00'
    
        wb.save(failo_pavadinimas)