import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from MatavimoVieta import Kaminas

def skaiciuoti_H2O2(kaminas_obj: Kaminas):
    failo_pavadinimas = 'Rezultatai.xlsx'
    lapas_pavadinimas = 'H2O'
    
    try:
        # Atidarome failą (visada naudokite pandas ir openpyxl)
        wb = load_workbook(failo_pavadinimas)
    except FileNotFoundError:
        print(f"Klaida: Failas {failo_pavadinimas} nerastas.")
        return

    if lapas_pavadinimas not in wb.sheetnames:
        print(f"Klaida: Lapas {lapas_pavadinimas} nerastas.")
        return

    ws = wb[lapas_pavadinimas]

    # --- K7 langelio pildymas ---
    cell_k7 = ws['K7']
    
    # Įrašome jūsų nurodytą formulę
    cell_k7.value = "=(100*((I7/(H7*J7))/0.8038))/(((I7/(H7*J7))/0.8038)+(1/J7))"
    
    # Formatavimas: centravimas ir 3 skaičiai po kablelio
    cell_k7.alignment = Alignment(horizontal='center', vertical='center')
    cell_k7.number_format = '0.000'

    # Išsaugome pakeitimus
    wb.save(failo_pavadinimas)