from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter  # Pridėta ši eilutė
import os

def irasyti_antrastes2(kaminas_obj):
    file = "Rezultatai.xlsx"
    
    if os.path.exists(file):
        wb = load_workbook(file)
    else:
        wb = Workbook()

    # Bendri stiliai
    thin_side = Side(style="thin")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_font = Font(bold=True)
    normal_font = Font(bold=False)

    def formatuoti_eilute(ws, eilute, stulpeliai, pradzios_col=1, plotis=14):
        for i, tekstas in enumerate(stulpeliai, pradzios_col):
            cell = ws.cell(row=eilute, column=i)
            cell.value = tekstas
            cell.alignment = center_wrap
            cell.border = border
            cell.font = normal_font
            ws.column_dimensions[get_column_letter(i)].width = plotis

    # --- 1. Lapas "H2O" ---
    ws_h2o = wb.create_sheet("H2O") if "H2O" not in wb.sheetnames else wb["H2O"]
    h2o_headers = ["", "Ėminių ėmimo laikas t, min.", "Paėmimo greitis, l/min", "Atmosferinis slėgis, hPa", "Išretėjimas, -hPa", "Temperatūra prie aspiratoriaus, Vtr oC", "Dujų tūris normaliomis sąlygomis, Vn l.", "Prasiurbtas dujų tūris normaliosiomis sąlygomis Vmn, Nm3", "Vandens kondensato masė m H2O, kg", "Sausų dujų tankis normaliosioms sąlygomis qn (kg/m3)", "Drėgmės kiekis X (%) (izokinetinei sistemai)", "Pastabos"]
    formatuoti_eilute(ws_h2o, 5, h2o_headers, plotis=14)
    ws_h2o.merge_cells(start_row=3, start_column=1, end_row=3, end_column=12)
    title_h2o = ws_h2o.cell(row=3, column=1)
    title_h2o.value = "Skaičiuoklė kietųjų dalelių koncentracijai nustatyti. H2O skaičiavimas"
    title_h2o.font = title_font
    title_h2o.alignment = center_wrap

    # --- 2. Lapas "Paėmimas" ---
   
    ws_paemimas = wb.create_sheet("Paėmimas") if "Paėmimas" not in wb.sheetnames else wb["Paėmimas"]
    
    # Nustatome pavadinimą ir reikšmę pagal kamino formą
    if kaminas_obj.forma == "A":
        ortakio_pav = "Ortakio skersmuo d, m"
        skersmuo_m = float(kaminas_obj.skersmuo) / 100
        # Suformuojame tekstą su dviem ženklais po kablelio ir pakeičiame tašką į kablelį
        ortakio_reiksme = f"{skersmuo_m:.2f}".replace('.', ',')
    else:
        ortakio_pav = "Ortakio matmenys, m"
        gylis_m = float(kaminas_obj.gylis) / 100
        plotis_m = float(kaminas_obj.plotis) / 100
        # Suformuojame tekstą formatu: gylis x plotis
        ortakio_reiksme = f"{gylis_m:.2f} x {plotis_m:.2f}".replace('.', ',')

    paemimas_headers = [
        "Matavimo data", "Ėminių registracijos Nr. T-107-2026-E-", "Objekto pavadinimas, adresas, taršos šaltinio Nr.",
        ortakio_pav, "Skerspjūvio plotas A (m2)", "Temperatūra ortakyje tor (°C)", "Vidutinis srauto greitis ortakyje wor (m/s)",
        "Atmosferinis slėgis P (hPa)", "Statinis slėgis ortakyje ΔPor (hPa)", "Slėgis prieš rotametrą ΔPr (±hPa)",
        "Antgalio diametras da (mm)", "Prasiurbtas dujų tūris Vm (m3)", "Siurbimo laikas t (min)", "Siurbimo greitis Vo (l/min)",
        "Siurbiamų dujų temperatūra prieš rotametrą tr (°C)", "O2iš (%) (ore 21 %)", "CO2iš (%) (ore 0,04 %)",
        "Darbuotojo inicialai, data", "Pastabos", "Išmatuota O2 koncentracija (%)", "Išmatuota CO2 koncentracija (%)", "Temperatūra ortakyje tor (°C)"
    ]
    
    # Įrašome antraštes (5 eilutė)
    formatuoti_eilute(ws_paemimas, 5, paemimas_headers, plotis=14)
    
    # Įrašome konkrečią reikšmę į 6 eilutę, 4 stulpelį (po ortakio pavadinimu)
    val_cell = ws_paemimas.cell(row=6, column=4)
    val_cell.value = ortakio_reiksme
    val_cell.alignment = center_wrap
    val_cell.border = Border()

    # --- 3. Lapas "Aerodinamika" ---
    ws_aero = wb.create_sheet("Aerodinamika") if "Aerodinamika" not in wb.sheetnames else wb["Aerodinamika"]
    aero_headers = ["Matavimo data", "Ėminių registracijos Nr. T-107-2026-E-", "Objekto pavadinimas, adresas, taršos šaltinio Nr.", "Prasiurbtas dujų tūris Vm (m3)", "Prasiurbtas dujų tūris normaliosiomis sąlygomis Vmn (Nm3)", "Prasiurbto dujų tūrio normaliosiomis sąlygomis suma Vmn(sum) (Nm3)", "Sausų dujų tankis normaliosioms sąlygomis qn (kg/m3)", "Drėgmės kiekis X (%)", "Pratrauktas drėgnų dujų tūris realiomis sąlygomis Vmk (m3)", "Dujų srauto greitis antgalyje v antg. (m/s)", "Izokinetiškumas (0,95-1,15)", "Izokinetiškumo vidurkis (0,95-1,15)", "Skaičiavimus atlikusio asmens inicialai, data", "Pastabos"]
    formatuoti_eilute(ws_aero, 5, aero_headers, plotis=14)
    ws_aero.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)
    title_a = ws_aero.cell(row=2, column=1)
    title_a.value = "Skaičiuoklė kietųjų dalelių koncentracijai nustatyti. Izokinetiškumo nustatymas"
    title_a.font = title_font
    title_a.alignment = center_wrap

    # --- 4. Lapas "Koncentracija" ---
    ws_konc = wb.create_sheet("Koncentracija") if "Koncentracija" not in wb.sheetnames else wb["Koncentracija"]
    konc_headers = ["Matavimo data", "Ėminių registracijos Nr. T-107-2026-E-", "Objekto pavadinimas, adresas, taršos šaltinio Nr.", "Filtro Nr.", "Filtro svoris prieš KD ėmimą mpr (g)", "Filtro svoris po KD ėmimo mpo (g)", "Tuščiojo ėminio KD svoris mt (g)", "Nuosėdų kiekis mn (g)", "Indų svorių pokytis mIt (g)", "Sausų dujų tūris n.s. Vmn (Nm3)", "Kietų dalelių koncentracija c1 (neįvert. O2) (mg/Nm3)", "Tuščiojo filtro vertė mft (mg/m3) (< 2 mg/m3)", "Surinktų dulkių ir tuščiojo ėminio santykis mFtk (>5)", "Skaičiavimus atlikusio asmens inicialai, data", "Filtrų skaičius Fsk (vnt.)", "Pastabos"]
    formatuoti_eilute(ws_konc, 5, konc_headers, plotis=14)
    ws_konc.merge_cells(start_row=2, start_column=1, end_row=2, end_column=16)
    title_k = ws_konc.cell(row=2, column=1)
    title_k.value = "Skaičiuoklė kietųjų dalelių koncentracijai nustatyti. Kietųjų dalelių koncentracijos nustatymas"
    title_k.font = title_font
    title_k.alignment = center_wrap

    # --- 5. Lapas "Svėrimas" ---
    ws_sver = wb.create_sheet("Svėrimas") if "Svėrimas" not in wb.sheetnames else wb["Svėrimas"]
    sver_col_width = 16
    for row_idx in [4, 5]:  # Antraščių eilutės
        ws_sver.row_dimensions[4].height = 60  # Leidžia Excel pačiam parinkti aukštį
    
    # Nustatome stulpelių plotį iš anksto
    for col_idx in range(1, 16):
        ws_sver.column_dimensions[get_column_letter(col_idx)].width = sver_col_width

    for col in [1, 2, 3, 4, 8, 9, 13, 14, 15]:
        ws_sver.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
    
    headers_sver_top = {
        1: "Ėminių registracijos Nr. T-107-2026-E-",
        2: "Objekto pavadinimas, adresas, taršos šaltinio Nr.",
        3: "Filtro/ indo Nr.",
        4: "Tuščio filtro/indo svėrimo data",
        8: "Atsakingo asmens inicialai, data",
        9: "Filtro/indo su ėminiu svėrimo data",
        13: "Kietųjų dalelių svoris mf/Svorių pokytis (g) (<2,20 g filtro talpumas)",
        14: "Atsakingo asmens inicialai, data",
        15: "Pastabos"
    }
    for col, txt in headers_sver_top.items():
        c = ws_sver.cell(row=4, column=col)
        c.value = txt
        c.alignment = center_wrap
        c.border = border
        ws_sver.cell(row=5, column=col).border = border

    # Grupiniai pavadinimai (E-G ir J-L)
    ws_sver.merge_cells(start_row=4, start_column=5, end_row=4, end_column=7)
    c_mpr = ws_sver.cell(row=4, column=5)
    c_mpr.value = "Filtro/indo svoris prieš KD ėmimą mpr (g)"
    c_mpr.alignment = center_wrap
    c_mpr.border = border
    for col in range(5, 8): 
        ws_sver.cell(row=4, column=col).border = border

    ws_sver.merge_cells(start_row=4, start_column=10, end_row=4, end_column=12)
    c_mpo = ws_sver.cell(row=4, column=10)
    c_mpo.value = "Filtro/indo svoris po KD ėmimo mpo (g)"
    c_mpo.alignment = center_wrap
    c_mpo.border = border
    for col in range(10, 13): 
        ws_sver.cell(row=4, column=col).border = border

    # 5 eilutės skaičiukai
    for i, col in enumerate(range(5, 8), 1):
        c = ws_sver.cell(row=5, column=col)
        c.value = str(i)
        c.alignment = center_wrap
        c.border = border
    for i, col in enumerate(range(10, 13), 1):
        c = ws_sver.cell(row=5, column=col)
        c.value = str(i)
        c.alignment = center_wrap
        c.border = border

    ws_sver.merge_cells(start_row=2, start_column=1, end_row=2, end_column=15)
    title_s = ws_sver.cell(row=2, column=1)
    title_s.value = "Skaičiuoklė kietųjų dalelių koncentracijai nustatyti. Svėrimas"
    title_s.font = title_font
    title_s.alignment = center_wrap

    # --- 6. Lapas "Koncentracijos ribinės vertės" ---
    ws_rib = wb.create_sheet("Koncentracijos ribinės vertės") if "Koncentracijos ribinės vertės" not in wb.sheetnames else wb["Koncentracijos ribinės vertės"]
    rib_headers = ["Planuojama surinkti dulkių masė m, g (<2,20 g)", "Mažiausia ėminio ėmimo trukmė t, 30 min", "Tikėtina dulkių koncentracija cexs, mg/Nm3;", "Planuojamas dujų tūrio siurbimo greitis matavimo metu ortakio sąlygomis Vo, l/min."]
    formatuoti_eilute(ws_rib, 5, rib_headers, plotis=18)
    ws_rib.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
    title_r = ws_rib.cell(row=3, column=1)
    title_r.value = "Patikrinimas, ar planuojamas surinkti dulkių kiekis suderinamas su tuščiojo ėminio verte ir ar filtras nebus per daug apkrautas"
    title_r.font = title_font
    title_r.alignment = center_wrap

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(file)