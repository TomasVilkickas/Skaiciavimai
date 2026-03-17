import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side

def sukurti_Paemimas_komplektus(kaminas_obj):
    """
    Paruošia struktūrą 'Pradiniai.xlsx' lape 'Paėmimas'.
    Sukuria koordinačių žodyną vėlesniam duomenų perkėlimui.
    """
    file_path = 'Pradiniai.xlsx'
    sheet_name = 'Paėmimas'
    
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])

    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]

    # Stiliai
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    antrastes = [
        "Slėgis prieš rotametrą ΔPr (±hPa)",
        "Antgalio diametras da (mm)",
        "Siurbimo laikas t (min)",
        "Siurbimo greitis Vo (l/min)",
        "Siurbiamų dujų temperatūra prieš rotametrą tr (°C)"
    ]

    # Stulpelių pločio nustatymas
    for col_idx in range(1, (kaminas_obj.liniju_skaicius * 6) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 18

    koordinaciu_zodynas = {}
    dabartine_eilute = 7
    master_diametras_coords = [] 
    master_laikas_coords = []

    for f_idx in range(1, kaminas_obj.filtru_skaicius + 1):
        ws.cell(row=dabartine_eilute, column=1, value=f"{f_idx} filtras").font = bold_font
        linijos_pradinis_stulpelis = 1
        
        for l_idx in range(1, kaminas_obj.liniju_skaicius + 1):
            # Linijos antraštė
            ws.merge_cells(start_row=dabartine_eilute + 1, start_column=linijos_pradinis_stulpelis, 
                           end_row=dabartine_eilute + 1, end_column=linijos_pradinis_stulpelis + 4)
            linijos_cell = ws.cell(row=dabartine_eilute + 1, column=linijos_pradinis_stulpelis, value=f"{l_idx} linija")
            linijos_cell.alignment = center_align
            linijos_cell.font = bold_font

            # Lentelės antraštės
            for col_offset, pavadinimas in enumerate(antrastes):
                cell = ws.cell(row=dabartine_eilute + 2, column=linijos_pradinis_stulpelis + col_offset, value=pavadinimas)
                cell.alignment = center_align
                cell.border = thin_border

            # ČIA BUVO KLAIDA: Sukuriamas sąrašas linijos taškams saugoti
            linijos_tasku_adresai = []
            
            for t_idx in range(1, kaminas_obj.tasku_skaicius + 1):
                r_idx = dabartine_eilute + 2 + t_idx
                vieno_tasko_stulpeliai = []
                
                for col_offset in range(5):
                    data_cell = ws.cell(row=r_idx, column=linijos_pradinis_stulpelis + col_offset)
                    data_cell.border = thin_border
                    data_cell.alignment = center_align
                    
                    # 1. Antgalio diametras
                    if col_offset == 1:
                        data_cell.number_format = '0.0'
                        if f_idx == 1 and l_idx == 1:
                            master_diametras_coords.append(data_cell.coordinate)
                        else:
                            data_cell.value = f"={master_diametras_coords[t_idx-1]}"
                    
                    # 2. Siurbimo laikas
                    elif col_offset == 2:
                        data_cell.number_format = '0'
                        if f_idx == 1 and l_idx == 1:
                            master_laikas_coords.append(data_cell.coordinate)
                        else:
                            data_cell.value = f"={master_laikas_coords[t_idx-1]}"
                    
                    # Formatai kitiems
                    elif col_offset in [0, 4]: data_cell.number_format = '0.0'
                    else: data_cell.number_format = '0'
                    
                    vieno_tasko_stulpeliai.append(data_cell.coordinate)
                
                # Pridedame šio taško visas 5 ląsteles į linijos sąrašą
                linijos_tasku_adresai.append(vieno_tasko_stulpeliai)
            
            # Įrašome visos linijos taškus į žodyną
            koordinaciu_zodynas[(f_idx, l_idx)] = linijos_tasku_adresai
            linijos_pradinis_stulpelis += 6

        dabartine_eilute += kaminas_obj.tasku_skaicius + 3 + 2

    wb.save(file_path)
    return koordinaciu_zodynas

def perkelti_paemimas_komplektus(kaminas_obj, koord_zodynas):
    pradiniai_path = 'Pradiniai.xlsx'
    rezultatai_path = 'Rezultatai.xlsx'
    sheet_name = 'Paėmimas'

    # Svarbu: data_only=True vis tiek paliekame dėl bendros praktikos
    wb_prad = load_workbook(pradiniai_path, data_only=True)
    if sheet_name not in wb_prad.sheetnames:
        return
    ws_prad = wb_prad[sheet_name]
    center_align = Alignment(horizontal='center', vertical='center')

    try:
        wb_rez = load_workbook(rezultatai_path)
    except FileNotFoundError:
        print(f"Klaida: {rezultatai_path} nerastas.")
        return

    if sheet_name not in wb_rez.sheetnames:
        wb_rez.create_sheet(sheet_name)
    ws_rez = wb_rez[sheet_name]

    headers = {}
    for r in range(1, 10):
        for col in range(1, ws_rez.max_column + 1):
            val = ws_rez.cell(row=r, column=col).value
            if val:
                headers[str(val).strip()] = col

    target_cols_list = [
        "Slėgis prieš rotametrą ΔPr (±hPa)",
        "Antgalio diametras da (mm)",
        "Siurbimo laikas t (min)",
        "Siurbimo greitis Vo (l/min)",
        "Siurbiamų dujų temperatūra prieš rotametrą tr (°C)"
    ]
    pastabos_col = headers.get("Pastabos")

    dabartine_rez_eilute = 6 

    # --- NAUJA DALIS: Sukuriame talpyklą (cache) pirmo filtro duomenims ---
    master_duomenys = {} # Saugojame: (taško_indeksas, stulpelio_indeksas) -> reikšmė

    for (f_idx, l_idx), tasku_sarasas in koord_zodynas.items():
        
        for t_idx, tasko_coords in enumerate(tasku_sarasas):
            for i, col_name in enumerate(target_cols_list):
                target_idx = headers.get(col_name)
                if target_idx:
                    # Nuskaitome reikšmę iš Excel
                    reiksme = ws_prad[tasko_coords[i]].value
                    
                    # LOGIKA: Jei tai 1 filtras ir 1 linija, išsisaugome reikšmes į atmintį
                    if f_idx == 1 and l_idx == 1:
                        master_duomenys[(t_idx, i)] = reiksme
                    
                    # LOGIKA: Antgalio diametras (i=1) ir Siurbimo laikas (i=2)
                    # Jei čia ne pirmas filtras/linija, imame reikšmę iš atminties, o ne iš Excel formulės
                    if (i == 1 or i == 2) and (f_idx != 1 or l_idx != 1):
                        reiksme = master_duomenys.get((t_idx, i))

                    cell = ws_rez.cell(row=dabartine_rez_eilute, column=target_idx, value=reiksme)
                    cell.alignment = center_align

                    # Tiksliname formatus: slėgis, antgalis ir tr (po kablelio)
                    if col_name in ["Slėgis prieš rotametrą ΔPr (±hPa)",
                                    "Antgalio diametras da (mm)",
                                    "Siurbiamų dujų temperatūra prieš rotametrą tr (°C)"]:
                        cell.number_format = '0.0'

            if pastabos_col and t_idx == 0:
                ws_rez.cell(row=dabartine_rez_eilute, column=pastabos_col).value = f"{f_idx} filtras, {l_idx} linija"            
            dabartine_rez_eilute += 1
        
        dabartine_rez_eilute += 4

    wb_rez.save(rezultatai_path)