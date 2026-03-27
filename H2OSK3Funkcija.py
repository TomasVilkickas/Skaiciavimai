import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from MatavimoVieta import Kaminas

def rasti_koordinates(ws, fraze):
    """Pagalbinė funkcija surasti langelio koordinates pagal tekstą."""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == fraze:
                return cell.row, cell.column
    return None, None

def skaiciuoti_H2O3(kaminas_obj: Kaminas):
    failo_pavadinimas = 'Rezultatai.xlsx'
    
    try:
        # Atidarome failą (visada naudokite pandas ir openpyxl) [cite: 2026-03-03]
        wb = load_workbook(failo_pavadinimas)
    except FileNotFoundError:
        print(f"Klaida: Failas {failo_pavadinimas} nerastas.")
        return

    # --- 1 DALIS: Aerodinamika (H stulpelis) ---
    if 'Aerodinamika' in wb.sheetnames and 'H2O' in wb.sheetnames:
        ws_aero = wb['Aerodinamika']
        t_skaicius = kaminas_obj.tasku_skaicius
        l_skaicius = kaminas_obj.liniju_skaicius
        f_skaicius = kaminas_obj.filtru_skaicius
        viso_lenteliu = l_skaicius * f_skaicius
        
        tikslo_eilute = 6 
        for _ in range(viso_lenteliu):
            for i in range(t_skaicius):
                cell_h = ws_aero.cell(row=tikslo_eilute, column=8) # H stulpelis
                cell_h.value = f"='H2O'!K7"
                cell_h.alignment = Alignment(horizontal='center', vertical='center')
                cell_h.number_format = '0.000'
                tikslo_eilute += 1
            tikslo_eilute += 4

    # --- 2 DALIS: Greitis (Kondensato masė ir Dujų tūris Vmn) ---
    if 'H2O' in wb.sheetnames and 'Greitis' in wb.sheetnames:
        ws_h2o = wb['H2O']
        ws_greitis = wb['Greitis']

        # Apskaičiuojame suvestinės eilutę pagal kaminas_obj parametrus
        t_skaicius = kaminas_obj.tasku_skaicius
        viso_lenteliu = kaminas_obj.liniju_skaicius * kaminas_obj.filtru_skaicius
        
        # Formulė: pradinė eilutė (6) + (lentelių skaičius * (taškų skaičius + 4 tarpai)) - 1
        suvestines_g_eilute = 6 + (viso_lenteliu * (t_skaicius + 4)) - 1
        
        # Užduotys: (Paieškos frazė, skaičių po kablelio kiekis)
        uzduotys = [
            ("Vandens kondensato masė m H2O, kg", '0.000'),
            ("Prasiurbtas dujų tūris normaliosiomis sąlygomis Vmn, Nm3", '0.000000')
        ]

        for fraze, formatas in uzduotys:
            h2o_row, h2o_col = rasti_koordinates(ws_h2o, fraze)
            greitis_row, greitis_col = rasti_koordinates(ws_greitis, fraze)

            if h2o_row and greitis_row:
                # Šaltinis: H2O lapas, antras langelis po fraze (+2)
                saltinio_langelis = f"{get_column_letter(h2o_col)}{h2o_row + 2}"
                
                # Tikslas: Greitis lapas, du langeliai po fraze (+1 ir +2)
                for offset in [1, 2]:
                    target_cell = ws_greitis.cell(row=greitis_row + offset, column=greitis_col)
                    target_cell.value = f"='H2O'!{saltinio_langelis}"
                    target_cell.alignment = Alignment(horizontal='center', vertical='center')
                    target_cell.number_format = formatas
            else:
                print(f"Pastaba: Frazė '{fraze}' nerasta.")

# --- C. Sausų dujų tankio perkėlimas iš Aerodinamika lapo ---
        tankio_fraze = "Sausų dujų tankis normaliosioms sąlygomis qn, (kg/m3)"
        gr_t_row, gr_t_col = rasti_koordinates(ws_greitis, tankio_fraze)

        if gr_t_row and 'Aerodinamika' in wb.sheetnames:
            # Rašome į vieną langelį žemyn po fraze (+1)
            cell_tankis = ws_greitis.cell(row=gr_t_row + 1, column=gr_t_col)
            
            # Nuoroda į Aerodinamika suvestinės langelį G stulpelyje
            # suvestines_g_eilute kintamasis turi būti apskaičiuotas 1-oje dalyje
            cell_tankis.value = f"='Aerodinamika'!G{suvestines_g_eilute}"
            
            # Formatavimas
            cell_tankis.alignment = Alignment(horizontal='center', vertical='center')
            cell_tankis.number_format = '0.00'

    # Išsaugome visus pakeitimus
    wb.save(failo_pavadinimas)