import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Border, Side
from MatavimoVieta import Kaminas

def spalvinti_paemimas2(ws, kaminas_obj, pradzios_eilute, yra_kopija=False):
    # Spalvos
    GELTONA = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    ZALIA = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    ORANZINE = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    MELYNA = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
    
    thin = Side(style='thin')
    all_sides = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    n = kaminas_obj.tasku_skaicius
    start_row = pradzios_eilute
    end_row = start_row + n - 1

    # --- 1. A, B, C STULPELIAI ---
    ws.cell(row=start_row, column=1).fill = ORANZINE # A6
    
    # B6 skirtumas: kopijoje oranžinė
    if yra_kopija:
        ws.cell(row=start_row, column=2).fill = ORANZINE # B6

    for i in range(n):
        ws.cell(row=start_row + i, column=3).fill = ORANZINE # C žemyn

    # --- 2. D ir E STULPELIAI (Skirtumas: kopijoje oranžinė vietoj žalios) ---
    de_spalva = ORANZINE if yra_kopija else ZALIA
    for col in [4, 5]:
        ws.cell(row=start_row, column=col).fill = de_spalva
        for r in range(start_row, end_row + 1):
            ws.cell(row=r, column=col).border = Border(
                left=thin, right=thin, 
                top=(thin if r == start_row else None), 
                bottom=(thin if r == end_row else None)
            )

    # --- 3. TINKLELIS NUO F IKI O ---
    for i in range(n):
        r = start_row + i
        ws.cell(row=r, column=6).fill = ORANZINE # F
        ws.cell(row=r, column=7).fill = GELTONA  # G
        ws.cell(row=r, column=8).fill = ORANZINE # H
        ws.cell(row=r, column=9).fill = ORANZINE # I
        ws.cell(row=r, column=10).fill = ZALIA   # J
        
        # K ir M (Skirtumas: kopijoje viskas oranžinė)
        k_m_fill = ORANZINE
        if i == 0 and not yra_kopija:
            k_m_fill = ZALIA
        ws.cell(row=r, column=11).fill = k_m_fill 
        ws.cell(row=r, column=13).fill = k_m_fill 
        
        ws.cell(row=r, column=12).fill = GELTONA  # L
        ws.cell(row=r, column=14).fill = ZALIA    # N
        ws.cell(row=r, column=15).fill = ZALIA    # O
        
        for col in range(6, 16):
            ws.cell(row=r, column=col).border = all_sides

    # --- 4. PAPILDOMI LANGELIAI APAČIOJE (L, M, N) ---
    extra_row_1 = end_row + 1
    extra_row_2 = end_row + 2
    tekstai = {12: "Suma Vm, m3", 13: "Suma t, min", 14: "Vidurkis Vo (l/min)"}
    for col, txt in tekstai.items():
        ws.cell(row=extra_row_1, column=col).fill = GELTONA
        ws.cell(row=extra_row_1, column=col).border = all_sides
        cell_txt = ws.cell(row=extra_row_2, column=col)
        cell_txt.value = txt
        cell_txt.alignment = center_alignment
        cell_txt.border = all_sides

    # --- 5. P, Q, R, S IR X, Y ---
    for i in range(n):
        r = start_row + i
        ws.cell(row=r, column=16).fill = ORANZINE # P
        ws.cell(row=r, column=17).fill = ORANZINE # Q
        ws.cell(row=r, column=16).border = all_sides
        ws.cell(row=r, column=17).border = all_sides
        
        # R6 skirtumas: kopijoje oranžinė
        if i == 0 and yra_kopija:
            ws.cell(row=r, column=18).fill = ORANZINE
        
        for col in [18, 19]: # R ir S išorinis rėmas
            ws.cell(row=r, column=col).border = Border(
                left=thin, right=thin, 
                top=(thin if r == start_row else None), 
                bottom=(thin if r == end_row else None)
            )

        ws.cell(row=r, column=24).fill = MELYNA   # X
        ws.cell(row=r, column=25).fill = ORANZINE # Y
        ws.cell(row=r, column=24).border = all_sides
        ws.cell(row=r, column=25).border = all_sides

        # --- 6. PAPILDOMAS BLOKELIS (T, U, V) TIK MOTININEI LENTELEI ---
    if not yra_kopija:
        # T=20, U=21, V=22
        tuv_tekstai = {20: "O2iš, (%) vidurkis", 21: "CO2iš, (%) vidurkis", 22: " tor vidurkis, °C"}
        
        for col in range(20, 23):
            # 6, 7, 8 eilutės (start_row, +1, +2) - ŽALIAI
            for r_idx in range(start_row, start_row + 3):
                cell = ws.cell(row=r_idx, column=col)
                cell.fill = ZALIA
                cell.border = all_sides
            
            # 9 eilutė (start_row + 3) - GELTONAI
            cell_9 = ws.cell(row=start_row + 3, column=col)
            cell_9.fill = GELTONA
            cell_9.border = all_sides
            
            # 10 eilutė (start_row + 4) - BE SPALVŲ + TEKSTAS
            cell_10 = ws.cell(row=start_row + 4, column=col)
            cell_10.value = tuv_tekstai[col]
            cell_10.alignment = center_alignment
            cell_10.border = all_sides