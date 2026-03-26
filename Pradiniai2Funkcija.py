import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

def nuskaityti_ir_perkelti_Greitis(kaminas_obj):
    """1. Sukuriamas tuščias rėmelis duomenų įvedimui faile Pradiniai.xlsx"""
    failo_pavadinimas = "Pradiniai.xlsx"
    lapas_pavadinimas = "Greitis"
    
    try:
        wb = load_workbook(failo_pavadinimas)
    except FileNotFoundError:
        return

    if lapas_pavadinimas not in wb.sheetnames:
        wb.create_sheet(lapas_pavadinimas)
    ws = wb[lapas_pavadinimas]

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold_font = Font(bold=True)

    liniju_sk = kaminas_obj.liniju_skaicius
    tasku_sk = kaminas_obj.tasku_skaicius

    if tasku_sk == 0:
        return

    # Pdi stulpelių kūrimas
    for l in range(1, liniju_sk + 1):
        col_idx = l
        pavadinimas = f"Išmatuotas diferencinis slėgis taškuose {l} linija, Pdi, hPa"
        
        cell_header = ws.cell(row=5, column=col_idx)
        cell_header.value = pavadinimas
        cell_header.font = bold_font
        cell_header.alignment = alignment
        cell_header.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

        for t in range(1, tasku_sk + 1):
            row_idx = 5 + t
            cell_data = ws.cell(row=row_idx, column=col_idx)
            cell_data.alignment = alignment
            cell_data.border = thin_border
            cell_data.number_format = '0.00'

    # Atmosferinio ir statinio slėgio stulpelių kūrimas
    stulpeliai_papildomi = [
        {"pav": "Atmosferinis slėgis P, hPa", "format": "0"},
        {"pav": "Statinis slėgis ortakyje ± ΔP, hPa", "format": "0.00"}
    ]

    for i, info in enumerate(stulpeliai_papildomi):
        col_idx = liniju_sk + 1 + i
        col_let = get_column_letter(col_idx)
        
        cell_header = ws.cell(row=5, column=col_idx)
        cell_header.value = info["pav"]
        cell_header.font = bold_font
        cell_header.alignment = alignment
        cell_header.border = thin_border
        ws.column_dimensions[col_let].width = 18

        pabaigos_eilute = 5 + tasku_sk
        for r in range(6, pabaigos_eilute + 1):
            ws.cell(row=r, column=col_idx).border = thin_border

        if tasku_sk > 1:
            ws.merge_cells(f"{col_let}6:{col_let}{pabaigos_eilute}")
        
        main_cell = ws.cell(row=6, column=col_idx)
        main_cell.alignment = alignment
        main_cell.number_format = info["format"]

        wb.save(failo_pavadinimas)

def paruosti_H2O_lapa():
    """Sukuriamas ir suformatuojamas H2O lapas faile Pradiniai.xlsx"""
    failo_pavadinimas = "Pradiniai.xlsx"
    lapas_pavadinimas = "H2O"
    
    try:
        wb = load_workbook(failo_pavadinimas)
    except FileNotFoundError:
        return

    if lapas_pavadinimas not in wb.sheetnames:
        wb.create_sheet(lapas_pavadinimas)
    ws = wb[lapas_pavadinimas]

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold_font = Font(bold=True)

    antrastes = [
        (2, "Ėminių ėmimo laikas t, min."),
        (3, "Paėmimo greitis, l/min"),
        (4, "Išretėjimas, -hPa"),
        (5, "Temperatūra prie aspiratoriaus, Vtr oC"),
        (6, "Vandens kondensato masė m H2O, kg")
    ]

    for col_idx, pavadinimas in antrastes:
        cell = ws.cell(row=5, column=col_idx)
        cell.value = pavadinimas
        cell.font = bold_font
        cell.alignment = alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    formatai = {2: "0", 3: "0", 4: "0.0", 5: "0.0", 6: "0.000"}

    for col_idx in range(2, 7):
        cell = ws.cell(row=6, column=col_idx)
        cell.border = thin_border
        cell.alignment = alignment
        if col_idx in formatai:
            cell.number_format = formatai[col_idx]
            
    wb.save(failo_pavadinimas)

def perkelti_greitis_duomenis(kaminas_obj):
    """Nuskaito duomenis iš Pradiniai.xlsx ir įrašo į Rezultatai.xlsx pagal pavadinimus"""
    failas_is = "Pradiniai.xlsx"
    failas_i = "Rezultatai.xlsx"
    lapas = "Greitis"
    
    # 1. Nuskaitymas naudojant pandas
    df_is = pd.read_excel(failas_is, sheet_name=lapas, header=4)
    
    # 2. Atidarymas įrašymui
    wb_rez = load_workbook(failas_i)
    ws_rez = wb_rez[lapas]
    centravimas = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Surandame lapą "Paėmimas" ir nustatome pradinę poziciją X stulpelyje (24 stulpelis)
    ws_paemimas = wb_rez["Paėmimas"]
    dabartine_eilute = 6
    x_stulpelis = 24  # X raidė atitinka 24-ą stulpelį

    # Surandame Rezultatai.xlsx stulpelių koordinates pagal pavadinimus
    rez_headers = {}
    for col in range(1, ws_rez.max_column + 1):
        header_val = ws_rez.cell(row=5, column=col).value
        if header_val:
            rez_headers[header_val] = col

    liniju_sk = kaminas_obj.liniju_skaicius
    tasku_sk = kaminas_obj.tasku_skaicius

    # 3. Perkeliame Pdi reikšmes
    for l in range(1, liniju_sk + 1):
        pav = f"Išmatuotas diferencinis slėgis taškuose {l} linija, Pdi, hPa"
        if pav in df_is.columns:
            if pav in rez_headers:
                col_dest = rez_headers[pav]
            for t_idx in range(tasku_sk):
                # .iloc[t_idx] paima reikšmę iš konkrečios eilutės
                reiksme = df_is[pav].iloc[t_idx]
                cell = ws_rez.cell(row=6 + t_idx, column=col_dest)
                cell.value = reiksme
                cell.alignment = centravimas
                cell.number_format = "0.00"

                # --- NAUJA: Perkėlimas į "Paėmimas" lapą į X stulpelį ---
            # --- Diferencinio slėgio perkėlimas į "Paėmimas" lapą ---
    x_stulpelis = 24  # X stulpelis
    dabartine_eilute = 6  # Pradinė eilutė lape "Paėmimas"

    # Pridedame ciklą per filtrų skaičių
    for f in range(kaminas_obj.filtru_skaicius):
        for l in range(1, liniju_sk + 1):
            pav = f"Išmatuotas diferencinis slėgis taškuose {l} linija, Pdi, hPa"
            
            if pav in df_is.columns:
                # Perrašome visus tos linijos taškus
                for t_idx in range(tasku_sk):
                    reiksme = df_is[pav].iloc[t_idx]
                    p_cell = ws_paemimas.cell(row=dabartine_eilute, column=x_stulpelis)
                    p_cell.value = reiksme
                    p_cell.alignment = centravimas
                    p_cell.number_format = "0.00"
                    dabartine_eilute += 1
                
                # Po kiekvienos linijos pridedame 4 eilučių tarpą (kaip buvo numatyta)
                dabartine_eilute += 4
        
        # Jei tarp filtrų blokų reikia papildomo tarpo arba specifinio eilučių valdymo, 
        # jį galima koreguoti čia. Šiuo metu ciklas tiesiog tęs pildymą žemyn.
                

    # 4. Perkeliame atmosferinį bei statinį slėgį (atkartojame per visus taškus)
    papildomi = ["Atmosferinis slėgis P, hPa", "Statinis slėgis ortakyje ± ΔP, hPa"]
    for pav in papildomi:
        if pav in df_is.columns and pav in rez_headers:
            col_dest = rez_headers[pav]
            # Pasiimame reikšmę iš pirmo duomenų langelio (nes Pradiniuose jis sulietas)
            reiksme_fiks = df_is[pav].iloc[0]
            for t_idx in range(tasku_sk):
                cell = ws_rez.cell(row=6 + t_idx, column=col_dest)
                cell.value = reiksme_fiks
                cell.alignment = centravimas

    wb_rez.save(failas_i)

def perkelti_H2O_duomenis():
    """Perkelia H2O duomenis iš Pradiniai.xlsx į Rezultatai.xlsx"""
    failas_is = "Pradiniai.xlsx"
    failas_i = "Rezultatai.xlsx"
    lapas = "H2O"

    try:
        wb_is = load_workbook(failas_is)
        ws_is = wb_is[lapas]
        wb_i = load_workbook(failas_i)
        ws_i = wb_i[lapas]
    except Exception as e:
        print(f"Klaida atidarant failus: {e}")
        return

    alignment = Alignment(horizontal="center", vertical="center")

    # Perkėlimo logika: (Iš_Cell, Į_Cell, Formatas)
    # B6 -> B7 (sveikas), C6 -> C7 (sveikas), D6 -> E7 (0.0), E6 -> F7 (0.0), F6 -> I7 (0.000)
    perkelimai = [
        ("B6", "B7", "0"),
        ("C6", "C7", "0"),
        ("D6", "E7", "0.0"),
        ("E6", "F7", "0.0"),
        ("F6", "I7", "0.000")
    ]

    for is_addr, i_addr, fmat in perkelimai:
        reiksme = ws_is[is_addr].value
        target_cell = ws_i[i_addr]
        target_cell.value = reiksme
        target_cell.alignment = alignment
        target_cell.number_format = fmat

    wb_i.save(failas_i)