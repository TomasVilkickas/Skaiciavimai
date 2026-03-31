import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from MatavimoVieta import Kaminas

def skaiciuoti_aerodinamika3(kaminas_obj):
    failo_pavadinimas = "Rezultatai.xlsx"
    lapo_pavadinimas = "Aerodinamika"
    
    try:
        wb = load_workbook(failo_pavadinimas)
    except FileNotFoundError:
        print(f"Klaida: Failas {failo_pavadinimas} nerastas.")
        return

    if lapo_pavadinimas not in wb.sheetnames:
        print(f"Klaida: Lapas '{lapo_pavadinimas}' nerastas.")
        return

    ws = wb[lapo_pavadinimas]
    
    tasku_skaicius = kaminas_obj.tasku_skaicius
    liniju_skaicius = kaminas_obj.liniju_skaicius
    filtru_skaicius = kaminas_obj.filtru_skaicius
    
    dabartine_eilute = 6
    tarpo_dydis = 4 

    # Einame per filtrus ir linijas
    for f in range(filtru_skaicius):
        for l in range(liniju_skaicius):
            
            # Lentelės pradžios eilutė vidurkiui skaičiuoti
            pradzios_eilute = dabartine_eilute
            
            # 1. Užpildome izokinetiškumo formules (K stulpelis)
            for t in range(tasku_skaicius):
                k_langelis = f"K{dabartine_eilute}"
                ws[k_langelis] = f"=J{dabartine_eilute}/Paėmimas!G{dabartine_eilute}"
                
                # Formatuojame K langelį
                ws[k_langelis].alignment = Alignment(horizontal="center", vertical="center")
                ws[k_langelis].number_format = "0.00"
                
                dabartine_eilute += 1
            
            # Paskutinė lentelės eilutė
            pabaigos_eilute = dabartine_eilute - 1
            
            # 2. Įrašome vidurkio formulę į L stulpelį (pirmąją lentelės eilutę)
            l_langelis = f"L{pradzios_eilute}"
            # Excel formulė: pvz., =AVERAGE(K6:K9)
            ws[l_langelis] = f"=AVERAGE(K{pradzios_eilute}:K{pabaigos_eilute})"
            
            # Formatuojame L langelį
            ws[l_langelis].alignment = Alignment(horizontal="center", vertical="center")
            ws[l_langelis].number_format = "0.00"
            
            # Pridedame tarpą tarp lentelių
            dabartine_eilute += tarpo_dydis

# --- Kodo dalis vidurkių vidurkiui (įterpti po ciklų) ---
    vidurkiu_langeliai = []
    seklys = 6 
    for f in range(filtru_skaicius):
        for l in range(liniju_skaicius):
            vidurkiu_langeliai.append(f"L{seklys}")
            seklys += tasku_skaicius + tarpo_dydis
    
    if vidurkiu_langeliai:
        # Atimame tarpo_dydį iš dabartine_eilute, kad panaikintume paskutinį tuščią tarpą
        galutinis_langelis = f"L{dabartine_eilute - tarpo_dydis}" 
        
        ws[galutinis_langelis] = f"=AVERAGE({','.join(vidurkiu_langeliai)})"
        
        # Formatavimas
        ws[galutinis_langelis].alignment = Alignment(horizontal="center", vertical="center")
        ws[galutinis_langelis].number_format = "0.00"
        
        try:
            from openpyxl.styles import Font
            ws[galutinis_langelis].font = Font(bold=True)
        except ImportError:
            pass
        
        wb.save(failo_pavadinimas)