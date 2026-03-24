import pandas as pd
import os
from MatavimoVieta import Kaminas
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font

def sukurti_sablona(kaminas_obj):
    """Sukuriamas Pradiniai.xlsx failas su suformatuota lentele duomenų įvedimui."""
    failo_pavadinimas = "Pradiniai.xlsx"
    lapas_pavadinimas = "Pradiniai"
    
    # Jei failas jau egzistuoja, jo neperrašome, kad neištrintume įvestų duomenų
    if os.path.exists(failo_pavadinimas):
        print(f"Pastaba: {failo_pavadinimas} jau egzistuoja, naujas šablonas nebus kuriamas.")
        return 

    # 1. Stulpelių pavadinimai
    stulpeliai = [
        "Matavimo data", 
        "Ėminių registracijos Nr. T-107-2026-E-", 
        "Objekto pavadinimas, adresas, taršos šaltinio Nr."
    ]
    
    # Sukuriame tuščią DataFrame su šiais stulpeliais
    df_template = pd.DataFrame(columns=stulpeliai)
    
    # 2. Įrašome į Excel naudojant pandas ir openpyxl [cite: 2026-03-03]
    with pd.ExcelWriter(failo_pavadinimas, engine='openpyxl') as writer:
        for pav in ["Pradiniai", "Greitis", "Paėmimas"]:
            if pav == lapas_pavadinimas:
                # Pradedame nuo 5-os eilutės (startrow=4)
                df_template.to_excel(writer, sheet_name=pav, startrow=4, index=False)
            else:
                pd.DataFrame().to_excel(writer, sheet_name=pav, index=False)
        
        # 3. Formatuojame išvaizdą (rėmeliai, BOLD, centravimas)
        ws = writer.sheets[lapas_pavadinimas]
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        bold_font = Font(bold=True)
        
        # Stulpelių pločiai
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 60
        
        # Eilučių aukščiai
        ws.row_dimensions[5].height = 40 # Antraštė
        ws.row_dimensions[6].height = 50 # Vieta įvedimui
        
        # Pritaikome stilius A, B ir C stulpeliams
        for col_num in range(1, 4):
            # Antraštės (5 eilutė)
            cell_header = ws.cell(row=5, column=col_num)
            cell_header.alignment = alignment
            cell_header.border = thin_border
            cell_header.font = bold_font
            
            # Įvedimo langeliai (6 eilutė)
            cell_input = ws.cell(row=6, column=col_num)
            cell_input.alignment = alignment
            cell_input.border = thin_border

    print(f"Sėkmingai sukurtas šablonas: {failo_pavadinimas}")

def nuskaityti_ir_perkelti_duomenis(kaminas_obj):
    """Nuskaito ranka įvestus duomenis iš Pradiniai.xlsx ir perkelia į Rezultatai.xlsx."""
    failo_pradiniai = "Pradiniai.xlsx"
    failo_rezultatai = "Rezultatai.xlsx"
    
    # 1. NUSKAITYMAS
    df = pd.read_excel(failo_pradiniai, sheet_name="Pradiniai", header=4)
    if df.empty:
        print("Klaida: Nerasta duomenų faile Pradiniai.xlsx!")
        return

    # Sutvarkome datą
    zalia_data = df.iloc[0, 0]
    if pd.api.types.is_datetime64_any_dtype(zalia_data) or not isinstance(zalia_data, str):
        tikra_data = pd.to_datetime(zalia_data).strftime('%Y-%m-%d')
    else:
        tikra_data = zalia_data

    duomenys = {
        'data': tikra_data,
        'reg_nr': df.iloc[0, 1],
        'objektas': df.iloc[0, 2]
    }

    # 2. ĮRAŠYMAS naudojant openpyxl [cite: 2026-03-03]
    wb = load_workbook(failo_rezultatai)
    
    # Apibrėžiame stilius, kad jie būtų pasiekiami visoje funkcijoje
    side_thin = Side(style='thin')
    thin_border = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)
    # Naujas lygiavimas: Horizontaliai centras, Vertikaliai viršus (top)
    top_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    # --- LAPAS "GREITIS" ---
    if "Greitis" in wb.sheetnames:
        ws_greitis = wb["Greitis"]
        for col_let, reiksme in {"A": duomenys['data'], "B": duomenys['reg_nr'], "C": duomenys['objektas']}.items():
            cell = ws_greitis[f"{col_let}6"]
            cell.value = reiksme
            cell.alignment = top_alignment

    # --- LAPAI "PAĖMIMAS" IR "AERODINAMIKA" ---
    # --- BENDRI DUOMENYS (Apibrėžiame čia, kad matytų visi lapai) ---
    perkeliami_duomenys = {"A": duomenys['data'], "B": duomenys['reg_nr'], "C": duomenys['objektas']}
    tasku_sk = kaminas_obj.tasku_skaicius
    tarpas = 4 # 4 tuščios eilutės tarp lentelių

    # --- LAPAI "PAĖMIMAS" IR "AERODINAMIKA" ---
    lapiu_sarasas = ["Paėmimas", "Aerodinamika"]

    for pavadinimas in lapiu_sarasas:
        if pavadinimas in wb.sheetnames:
            ws_temp = wb[pavadinimas]
            dabartine_eilute = 6 # Pirma lentelė prasideda 6-oje eilutėje

            # Sukame ciklus per filtrus ir linijas (iš MatavimoVieta.py)
            for f in range(kaminas_obj.filtru_skaicius):
                for l in range(kaminas_obj.liniju_skaicius):
                    pabaigos_eilute = dabartine_eilute + tasku_sk - 1
                    
                    for col_let, reiksme in perkeliami_duomenys.items():
                        col_idx = 1 if col_let == "A" else (2 if col_let == "B" else 3)
                        
                        # Įrašome reikšmę į viršutinį lentelės langelį
                        cell = ws_temp.cell(row=dabartine_eilute, column=col_idx)
                        cell.value = reiksme
                        cell.alignment = top_alignment

                        if col_let in ["A", "B"]:
                            # A ir B: Tik išorinis rėmas visam stulpelio blokui
                            for r in range(dabartine_eilute, pabaigos_eilute + 1):
                                current_cell = ws_temp.cell(row=r, column=col_idx)
                                current_cell.border = Border(
                                    left=side_thin, 
                                    right=side_thin, 
                                    top=side_thin if r == dabartine_eilute else None, 
                                    bottom=side_thin if r == pabaigos_eilute else None
                                )
                        else:
                            # C stulpelis: Suliejame ir uždedame pilną rėmą
                            for r in range(dabartine_eilute, pabaigos_eilute + 1):
                                ws_temp.cell(row=r, column=col_idx).border = thin_border
                            
                            if tasku_sk > 1:
                                ws_temp.merge_cells(start_row=dabartine_eilute, start_column=col_idx, 
                                                    end_row=pabaigos_eilute, end_column=col_idx)

                    # Perskaičiuojame poziciją kitai lentelei (taškai + 2 papildomos eilutės + tarpas)
                    dabartine_eilute += tasku_sk + tarpas

    # --- LAPAS "KONCENTRACIJA" ---
    if "Koncentracija" in wb.sheetnames:
        ws_konc = wb["Koncentracija"]
        
        # 1. Pirmiausia panaikiname senus suliejimus, jei jie egzistuoja (kad išvengtume klaidų) [cite: 2026-03-03]
        for range_ in list(ws_konc.merged_cells.ranges):
            if "C6" in range_ or "C7" in range_ or "C8" in range_:
                ws_konc.unmerge_cells(str(range_))

        # 2. Nustatome eilučių aukštį
        for r_idx in range(6, 9):
            ws_konc.row_dimensions[r_idx].height = 25
            
        for col_let, reiksme in perkeliami_duomenys.items():
            col_idx = list(perkeliami_duomenys.keys()).index(col_let) + 1
            
            if col_let in ["A", "B"]:
                # A ir B: Lygiavimas viršuje, be suliejimo, tik išorinis rėmas [cite: 2026-03-03]
                for r_idx in range(6, 9):
                    cell = ws_konc.cell(row=r_idx, column=col_idx)
                    if r_idx == 6:
                        cell.value = reiksme
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                    cell.border = Border(
                        left=side_thin, 
                        right=side_thin, 
                        top=side_thin if r_idx == 6 else None, 
                        bottom=side_thin if r_idx == 8 else None
                    )
            elif col_let == "C":
                # --- C STULPELIS: SULIEJAMAS, LYGIAVIMAS VIRŠUJE IR CENTRE ---
                cell_c = ws_konc.cell(row=6, column=3)
                cell_c.value = reiksme
                
                # SULIEJAME C6:C8 [cite: 2026-03-03]
                ws_konc.merge_cells("C6:C8")
                
                # Nustatome stilių sulietam plotui
                # Kadangi tai C stulpelis, nustatome horizontal="center" ir vertical="top"
                cell_c.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                
                # Uždėdami rėmelius sulietam plotui, openpyxl reikalauja, kad rėmelis būtų visoms ląstelėms
                for r_idx in range(6, 9):
                    ws_konc.cell(row=r_idx, column=3).border = thin_border
                    
    # --- LAPAS "SVĖRIMAS" ---
    if "Svėrimas" in wb.sheetnames:
        ws_sverimas = wb["Svėrimas"]
        
        # 1. Paruošiame duomenis abiem stulpeliams
        sverimo_duomenys = {"A": duomenys['reg_nr'], "B": duomenys['objektas']}
        s_row, e_row = 6, 16 # Duomenų rėžis nuo 6 iki 16 eilutės

        for col_let, reiksme in sverimo_duomenys.items():
            col_idx = 1 if col_let == "A" else 2
            
            # Įrašome reikšmę į viršutinį langelį
            ws_sverimas[f"{col_let}6"].value = reiksme
            
            # Panaikiname senus suliejimus, kad perrašytume švariai [cite: 2026-03-03]
            for range_ in list(ws_sverimas.merged_cells.ranges):
                if f"{col_let}6" in range_:
                    ws_sverimas.unmerge_cells(str(range_))

            if col_let == "A":
                # --- A STULPELIS: BE SULIEJIMO, TIK IŠORINIS RĖMAS ---
                for r_idx in range(s_row, e_row + 1):
                    cell = ws_sverimas.cell(row=r_idx, column=col_idx)
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                    # Rėmelis tik bloko išorei
                    cell.border = Border(
                        left=side_thin, 
                        right=side_thin, 
                        top=side_thin if r_idx == s_row else None, 
                        bottom=side_thin if r_idx == e_row else None
                    )
            else:
                # --- B STULPELIS: SULIETAS, LYGIAVIMAS VIRŠUJE ---
                # Suliejame B6:B16
                ws_sverimas.merge_cells(start_row=s_row, start_column=col_idx, end_row=e_row, end_column=col_idx)
                
                # Nustatome stilių visoms ląstelėms bloke (dėl rėmelių ir lygiavimo) [cite: 2026-03-03]
                for r_idx in range(s_row, e_row + 1):
                    cell = ws_sverimas.cell(row=r_idx, column=col_idx)
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                    cell.border = thin_border # B stulpeliui paliekame pilną rėmą aplink sulietą plotą

    wb.save(failo_rezultatai)