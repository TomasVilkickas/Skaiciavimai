import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from MatavimoVieta import Kaminas
from AerodinamikaSP2Funkcija import spalvinti_aerodinamika2

def spalvinti_aerodinamika1(kaminas_obj: Kaminas):
    file_path = "Rezultatai.xlsx"
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()

    ws = wb["Aerodinamika"] if "Aerodinamika" in wb.sheetnames else wb.create_sheet("Aerodinamika")

    dabartine_eilute = 6
    tarpas = 4 # Galite pasikeisti į 2 ar 3, jei norite mažesnio tarpo

    # Sukame ciklus: Filtrai x Linijos
    for f in range(kaminas_obj.filtru_skaicius):
        for l in range(kaminas_obj.liniju_skaicius):
            
            spalvinti_aerodinamika2(ws, kaminas_obj, dabartine_eilute)
            
            # Poslinkis: taškai + tarpas (čia aerodinamikoje nėra papildomų eilučių sumoms)
            dabartine_eilute += kaminas_obj.tasku_skaicius + tarpas

# --- SUVESTINĖS DALIS LAPO PABAIGOJE ---
    GELTONA = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    thin = Side(style='thin')
    all_sides = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold_font = Font(bold=True)
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)

    dabartine_eilute -= 4
    # 1. F IR L STULPELIAI (Iškart po paskutinės lentelės)
    # F (6 stulpelis), L (12 stulpelis)
    suvestine_fl = {
        6: {"tekstas": "Viso suma Vmn(sum), Nm3.", "formatas": "0.000000"},
        12: {"tekstas": "Izokinetiškumo vidurkių vidurkis", "formatas": "0.00"}
    }

    for col, nustatymai in suvestine_fl.items():
        # Rezultato langelis
        res_cell = ws.cell(row=dabartine_eilute, column=col)
        res_cell.fill = GELTONA
        res_cell.border = all_sides
        res_cell.font = bold_font
        res_cell.alignment = center_wrap
        res_cell.number_format = nustatymai["formatas"]
        
        # Teksto langelis žemiau
        txt_cell = ws.cell(row=dabartine_eilute + 1, column=col)
        txt_cell.value = nustatymai["tekstas"]
        txt_cell.border = all_sides
        txt_cell.alignment = center_wrap

    # 2. G IR I STULPELIAI (Praleidus 3 eilutes po paskutinės lentelės)
    # G (7 stulpelis), I (9 stulpelis)
    dabartine_eilute += 3 
    suvestine_gi = {
        7: {"tekstas": "Vidurkių vidurkis qn, (kg/m3)", "formatas": "0.00"},
        9: {"tekstas": "Viso suma Vmk(sum), m3.", "formatas": "0.000"}
    }

    for col, nustatymai in suvestine_gi.items():
        # Rezultato langelis
        res_cell = ws.cell(row=dabartine_eilute, column=col)
        res_cell.fill = GELTONA
        res_cell.border = all_sides
        res_cell.font = bold_font
        res_cell.alignment = center_wrap
        res_cell.number_format = nustatymai["formatas"]
        
        # Teksto langelis žemiau
        txt_cell = ws.cell(row=dabartine_eilute + 1, column=col)
        txt_cell.value = nustatymai["tekstas"]
        txt_cell.border = all_sides
        txt_cell.alignment = center_wrap
    wb.save(file_path)