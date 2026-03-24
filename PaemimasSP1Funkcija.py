from openpyxl import load_workbook
from MatavimoVieta import Kaminas
from PaemimasSP2Funkcija import spalvinti_paemimas2

def spalvinti_paemimas1(kaminas_obj: Kaminas):
    file_path = "Rezultatai.xlsx"
    
    # 1. Atidarome failą
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        from openpyxl import Workbook
        wb = Workbook()

    # 2. APIBRĖŽIAME 'ws' kintamąjį (kad išvengtume UnboundLocalError)
    if "Paėmimas" in wb.sheetnames:
        ws = wb["Paėmimas"]
    else:
        ws = wb.create_sheet("Paėmimas")

    dabartine_eilute = 6
    tarpas = 4
    ar_pirma_lentele = True

    # 3. Sukame ciklus
    for f in range(kaminas_obj.filtru_skaicius):
        for l in range(kaminas_obj.liniju_skaicius):
            
            yra_kopija = not ar_pirma_lentele
            
            # Kviečiame darbininką, dabar 'ws' jau tikrai egzistuoja
            spalvinti_paemimas2(
                ws=ws, 
                kaminas_obj=kaminas_obj, 
                pradzios_eilute=dabartine_eilute, 
                yra_kopija=yra_kopija
            )
            
            # Perstumiame eilutę žemyn
            dabartine_eilute += kaminas_obj.tasku_skaicius + tarpas
            ar_pirma_lentele = False

# --- SUVESTINĖS DALIS LAPO PABAIGOJE (M IR N STULPELIAI) ---
    # Grįžtame per tarpą atgal, kad būtume prie paskutinės lentelės rėmo
    dabartine_eilute -= 2 
    
    # Praleidžiame vieną eilutę po lentelės (kaip prašėte)
    suvestines_row = dabartine_eilute + 1 
    
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    GELTONA = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    thin = Side(style='thin')
    all_sides = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold_font = Font(bold=True)
    center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Konfigūracija M (13) ir N (14) stulpeliams
    suvestine_mn = {
        13: "Viso suma t, min",
        14: "Vidurkių vidurkis Vo vid (l/min) "
    }

    for col, tekstas in suvestine_mn.items():
        # 1. Rezultato langelis (Geltonas, Bold, Sveikas skaičius)
        res_cell = ws.cell(row=suvestines_row, column=col)
        res_cell.fill = GELTONA
        res_cell.border = all_sides
        res_cell.font = bold_font
        res_cell.alignment = center_wrap
        res_cell.number_format = '0'  # Sveikas skaičius be kablelių
        
        # 2. Teksto langelis žemiau (Be spalvos, Aprėmintas)
        txt_cell = ws.cell(row=suvestines_row + 1, column=col)
        txt_cell.value = tekstas
        txt_cell.border = all_sides
        txt_cell.alignment = center_wrap
    # 4. Išsaugojame
    wb.save(file_path)