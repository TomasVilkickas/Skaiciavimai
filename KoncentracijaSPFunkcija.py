import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill
from MatavimoVieta import Kaminas

def spalvinti_koncentracija(kaminas_obj: Kaminas):
    file_path = "Rezultatai.xlsx"
    sheet_name = "Koncentracija"
    
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]

    # Stiliai
    thin_side = Side(border_style="thin", color="000000")
    oranzine_uzpilda = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    geltona_uzpilda = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    centruoti = Alignment(horizontal="center", vertical="center")
    kairėje = Alignment(horizontal="left", vertical="center")

    # Reikalingas eilučių skaičius pagal filtrus
    filtru_sk = max(1, min(kaminas_obj.filtru_skaicius, 3))
    duomenu_range = range(6, 6 + filtru_sk)

    # 1. Nekintama dalis: A6, B6, C6 (C6, C7, C8 sulieti)
    for col in ['A', 'B']:
        cell = ws[f"{col}6"]
        cell.fill = oranzine_uzpilda
        cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

    # C stulpelio suliejimo spalvinimas
    ws.cell(row=6, column=3).fill = oranzine_uzpilda

    # 2. Dinaminis spalvinimas pagal filtrų skaičių (D-K stulpeliai)
    oranziniai_cols = [4, 5, 6, 7, 9, 10] # D, E, F, G, I, J
    geltoni_cols = [8, 11]               # H, K

    for r in duomenu_range:
        for c in range(4, 12): # D iki K
            cell = ws.cell(row=r, column=c)
            cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
            if c in oranziniai_cols:
                cell.fill = oranzine_uzpilda
            elif c in geltoni_cols:
                cell.fill = geltona_uzpilda

    # 3. L stulpelis: Panaikinti suliejimą, spalvinti geltonai
    for range_ in list(ws.merged_cells.ranges):
        if "L6" in range_:
            ws.unmerge_cells(str(range_))
    
    for r in duomenu_range:
        cell = ws.cell(row=r, column=12)
        cell.fill = geltona_uzpilda
        cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

    # 4. M stulpelis: Sulietas (M6:M8), geltonas
    for range_ in list(ws.merged_cells.ranges):
        if "M6" in range_:
            ws.unmerge_cells(str(range_))
    ws.merge_cells(start_row=6, start_column=13, end_row=8, end_column=13)
    m_cell = ws.cell(row=6, column=13)
    m_cell.fill = geltona_uzpilda
    m_cell.alignment = centruoti
    m_cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

    # N stulpelis: spalvinama tik 6 eilutė (oranžinė)
    # 1. Spalviname tik N6
    ws["N6"].fill = oranzine_uzpilda

    # 2. Sukuriame išorinį rėmą blokui N6:N8
    top_side = Side(border_style="thin", color="000000")
    bottom_side = Side(border_style="thin", color="000000")
    left_side = Side(border_style="thin", color="000000")
    right_side = Side(border_style="thin", color="000000")

    # Viršutinis langelis (N6)
    ws["N6"].border = Border(top=top_side, left=left_side, right=right_side)

    # Vidurinis langelis (N7)
    ws["N7"].border = Border(left=left_side, right=right_side)

    # Apatinis langelis (N8)
    ws["N8"].border = Border(bottom=bottom_side, left=left_side, right=right_side)

    # 5. O, P stulpeliai
    for r in duomenu_range:
        
        # O stulpelis (Be spalvos, tik rėmelis ir reikšmė 6-oje eilutėje)
        o_cell = ws.cell(row=r, column=15)
        o_cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
        if r == 6:
            o_cell.value = kaminas_obj.filtru_skaicius
            o_cell.alignment = centruoti

        # P stulpelis (Oranžinė + Tekstas)
        p_cell = ws.cell(row=r, column=16)
        p_cell.fill = oranzine_uzpilda
        p_cell.border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

    # P stulpelio tekstas ir pločio nustatymas
    lin_tekstas = "linija" if kaminas_obj.liniju_skaicius == 1 else "linijos"
    for i in range(filtru_sk):
        target_row = 6 + i
        ws.cell(row=target_row, column=16).value = f"{kaminas_obj.liniju_skaicius} {lin_tekstas}, {i+1} filtras"
        ws.cell(row=target_row, column=16).alignment = kairėje

    ws.column_dimensions['P'].width = 17 # Praplatiname P stulpelį

    wb.save(file_path)