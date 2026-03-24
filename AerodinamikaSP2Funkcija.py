import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from MatavimoVieta import Kaminas

def spalvinti_aerodinamika2(ws, kaminas_obj, pradzios_eilute):
    # Spalvos
    GELTONA = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    ORANZINE = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    
    thin = Side(style='thin')
    all_sides = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    n = kaminas_obj.tasku_skaicius
    start_row = pradzios_eilute
    end_row = start_row + n - 1

    # 1. A, B, C stulpeliai (tik pirma eilutė - oranžinė)
    for col in range(1, 4): # A, B, C
        ws.cell(row=start_row, column=col).fill = ORANZINE

    # 2. D - Oranžinė (kiekvienas langelis su rėmeliu)
    for i in range(n):
        cell = ws.cell(row=start_row + i, column=4)
        cell.fill = ORANZINE
        cell.border = all_sides

    # 3. E ir G - Geltona (kiekvienas langelis su rėmeliu)
    for col in [5, 7]:
        for i in range(n):
            cell = ws.cell(row=start_row + i, column=col)
            cell.fill = GELTONA
            cell.border = all_sides

    # --- PAPILDOMI PRIEDAI G IR I STULPELIAMS (KIEKVIENAI LENTELEI) ---
    for col_idx, tekstas in {7: "Vidurkis qn, (kg/m3)", 9: "Suma Vmk (m3)"}.items():
        # 1. Langelis skaičiui (end_row + 1)
        res_cell = ws.cell(row=end_row + 1, column=col_idx)
        res_cell.fill = GELTONA
        res_cell.border = all_sides
        res_cell.font = Font(bold=True)
        # Jei meta klaidą dėl center_alignment, naudojame tiesioginį aprašymą:
        res_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        res_cell.number_format = '0.00'
        
        # 2. Langelis tekstui (end_row + 2)
        txt_cell = ws.cell(row=end_row + 2, column=col_idx)
        txt_cell.value = tekstas
        txt_cell.border = all_sides
        txt_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 4. F - Geltona (TIK pirma eilutė, bet bendras rėmas)
    ws.cell(row=start_row, column=6).fill = GELTONA
    # Bendro rėmo logika stulpeliui F
    ws.cell(row=start_row, column=6).border = Border(top=thin, left=thin, right=thin)
    ws.cell(row=end_row, column=6).border = Border(bottom=thin, left=thin, right=thin)
    for r in range(start_row + 1, end_row):
        ws.cell(row=r, column=6).border = Border(left=thin, right=thin)

    # 5. H - Oranžinė (kiekvienas langelis su rėmeliu)
    for i in range(n):
        cell = ws.cell(row=start_row + i, column=8)
        cell.fill = ORANZINE
        cell.border = all_sides

    # 6. I, J, K - Geltona (kiekvienas langelis su rėmeliu)
    for col in [9, 10, 11]:
        for i in range(n):
            cell = ws.cell(row=start_row + i, column=col)
            cell.fill = GELTONA
            cell.border = all_sides

    # 7. L (Geltona), M (Oranžinė), N (Oranžinė) - TIK pirma eilutė nuspalvinta
    stulpeliu_konfig = {12: GELTONA, 13: ORANZINE, 14: ORANZINE}
    for col, spalva in stulpeliu_konfig.items():
        # Nuspalviname tik pirmą langelį
        ws.cell(row=start_row, column=col).fill = spalva
        
        # Bendro rėmo logika stulpeliams L, M, N
        ws.cell(row=start_row, column=col).border = Border(top=thin, left=thin, right=thin)
        ws.cell(row=end_row, column=col).border = Border(bottom=thin, left=thin, right=thin)
        for r in range(start_row + 1, end_row):
            ws.cell(row=r, column=col).border = Border(left=thin, right=thin)